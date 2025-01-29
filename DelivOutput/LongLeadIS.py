import re
from openpyxl import load_workbook
from tkinter import filedialog, messagebox
import os

class LongLeadGenerator:
    def __init__(self, workbook, wo_number, county, city_place):
        self.source_workbook = workbook
        self.wo_number = wo_number
        self.county = county
        self.city_place = city_place
        self.template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'longLeadTemplate.xlsx')
        
        # Original material_data dictionary remains the same as in your code
        self.material_data = {
            # Fiberglass Poles
            'FG Round 30\' CL5 TU440': ('1008200', 2034.68),
            'FG Round 35\' CL4 TU440': ('1008201', 1893.62),
            'FG Round 40\' CL3 TU440': ('1008202', 1802.70),
            'FG Round 45\' CL2 TU440': ('1008204', 2141.52),
            'FG Round 45\' CL1 TU450': ('8006086', 4664.65),
            'FG Round 50\' CL2 TU440': ('8005888', 5323.65),
            'FG Round 50\' CL1 TU450': ('8005890', 3756.91),
            'FG Round 55\' CL3 TU440': ('8005875', 3778.01),
            'FG Round 55\' CL1 TU450': ('8005891', 5757.24),
            'FG Round 60\' CL2 TU440': ('8005889', 6136.61),
            'FG Round 60\' CL1 TU450': ('8005892', 6136.61),

            # RSP Steel Poles
            'RSP-45-345 (RSP 70\')': ('8005031', 9035.64),
            'RSP-45-567 (RSP 60\')': ('8005015', 10006.71),
            'RSP-50-567 (RSP 60\')': ('8005015', 10006.71),
            'RSP-55-4567 (RSP 60\')': ('8005015', 10006.71),
            'RSP-60-4567': ('8005015', 10006.71),
            'RSP-60-6789': ('8005021', 11786.52),
            'RSP-60-891011 (RSP 80\')': ('8005040', 19399.88),
            'RSP-65-4567': ('8005024', 8615.86),
            'RSP-65-6789': ('8005029', 11786.52),
            'RSP-65-891011 (RSP 80\')': ('8005040', 19399.88),
            'RSP-70-34567': ('8005031', 9035.64),
            'RSP-70-56789': ('8005035', 13579.31),
            'RSP-75-34567': ('8005036', 9035.64),
            'RSP-75-56789': ('8005037', 13579.31),
            'RSP-80-34567 (RSP 70\')': ('8005031', 9035.64),
            'RSP-80-56789': ('8005039', 13579.31),
            'RSP-80-67891011': ('8005040', 19399.88),
            'RSP-85-456789': ('8005043', 14473.75),
            'RSP-85-67891011': ('8005044', 19399.88),
            'RSP-90-456789': ('8005045', 14473.75),
            'RSP-90-67891011': ('8005046', 19399.88),
            'RSP-95-456789': ('8005048', 14473.75),
            'RSP-95-67891011': ('8005051', 19399.88),
            'RSP-100-3456789': ('8005055', 17174.88),
            'RSP-100-567891011': ('8005057', 21192.67),
            'RSP-105-3456789 (RSP 100\')': ('8005055', 17174.88),
            'RSP-105-567891011': ('8005058', 21192.67),
            'RSP-110-3456789 (RSP 100\')': ('8005055', 17174.88),
            'RSP-110-567891011': ('8005061', 20736.55),
            'RSP-115-4567891011': ('8005062', 23832.09),
            'RSP-120-4567891011': ('8005063', 23832.09),

            # Hardware and Components
            'Pin Insulator': ('7999666', 32.38),
            'DE Insulator': ('5840285', 17.05),
            '2PO Tan xarm': ('7999557', 131.77),
            '2PO DE xarm': ('7999723', 258.25),
            '4PO Tan xarm': ('7999558', 196.43),
            '4PO DE xarm': ('7999724', 390.26),
            'Double fiberglass crossarm - 10\'': ('8006487', 1628.51),
            'Deadend cover': ('8006194', 56.88),
            'Stirrup cover': ('8006513', 43.03),
            'Connector cover (large)': ('8006442', 33.53),
            'Connector cover (small)': ('8006443', 32.22),

            # Tree Wire and ACSR
            '1/0 ACSR 15kV Tree Wire': ('8006173', 1.59),
            '4/0 ACSR 15kV Tree Wire': ('8006174', 2.55),
            '477 ACSR 15kV Tree Wire': ('8006175', 7.40),
            '795 ACSR 15kV Tree Wire': ('8006177', 6.35),
            '1/0 ACSR Raven (Tree Wire Neutral)': ('4508701', 0.51),
            '3/0 ACSR Pigeon (Tree Wire Neutral)': ('4508909', 0.37),
            '4/0 ACSR Penguin (Tree Wire Neutral)': ('4509006', 0.59),
            '477 AAC Cosmos (Tree Wire Neutral)': ('4505077', 1.49),

            # AAC Conductors
            '1/0 AAC, 15kV, grey, Covered Spacer Conductor': ('8005893', 0.88),
            '4/0 AAC, 15kV, grey, Covered Spacer Conductor': ('8005894', 1.43),
            '477 AAC, 15kV, grey, Covered Spacer Conductor': ('8005895', 5.18),
            '795 AAC, 15kV, grey, Covered Spacer Conductor': ('8005896', 4.39),
            '1/0 AAC, 35kV, grey, Covered Spacer Conductor': ('8005897', 1.41),
            '4/0 AAC, 35kV, grey, Covered Spacer Conductor': ('8005898', 7.74),
            '477 AAC, 35kV, grey, Covered Spacer Conductor': ('8005899', 3.61),
            '795 AAC, 35kV, grey, Covered Spacer Conductor': ('8005900', 5.56),

            # Messenger Wire
            '1/0 AWA Messenger': ('8004771', 1.43),
            '4/0 AWA Messenger': ('8004772', 4.09),

            # Guy Wire and Anchors
            '250\' coil of 7/16" UG guy wire': ('6155402', 212.50),
            '14" Screw Anchor': ('6107247', 72.95),
            '24" plate anchor': ('6101992', 222.69),
            '30" Plate Anchor': ('6102008', 411.45),

            # Transmission Components
            'TRANSMISSION: INSUL,POST,46/69KV': ('7992868', 160.50),
            'TRANSMISSION: INSUL,POST,46/69KV': ('7992873', 214.00),
            'TRANSMISSION: INSUL,POST,69KV,VE': ('5800313', 93.29),
            'TRANSMISSION: INSUL,SUSP,POLY,11': ('5800340', 60.14),
            'TRANSMISSION: XARM,STL,10\'-0",46': ('1600201', 735.40),
            'TRANSMISSION: HIGH STRENGTH INSUL,POST,46/69KV': ('7992863', 191.71),
            'TRANSMISSION: INSUL, POST, HORIZ, 49/69KV': ('5800347', 175.09),
            'TRANSMISSION: HS POLY 115kV (Horiz.)': ('7992864', 272.41),
            'TRANSMISSION: POLY 115kV (Horiz.)': ('5800348', 220.48),
            'TRANSMISSION: VERT POLY 115kV': ('5800354', 201.04),
            'TRANSMISSION: 115KV CROSSARM': ('1600204', 898.88),
            'TRANSMISSION: HS POLY 161kV (Horiz. Drop Eye Attachment, TD831K)': ('7992866', 306.00),
            'TRANSMISSION: HS POLY 161kV (Horiz. Trunnion Clamp, TD831Y)': ('8003013', 1063.35),
            'TRANSMISSION: POLY 161kV (Horiz.)': ('5800350', 489.00),
            'TRANSMISSION: POLY, SUSP, 161kV': ('5800343', 148.21),

            # Wood Poles
            'Pole,wood,35 Ft,class 4': ('1614354', 430.60),
            'Pole,wood,40 Ft,class 1': ('1614401', 745.89),
            'Pole,wood,45 Ft,class 1': ('1614451', 982.38),
            'Pole,wood,45 Ft,class 2': ('1614452', 989.48),
            'Pole,wood,45 Ft,class 3': ('1614453', 810.41),
            'Pole,wood,50 Ft,class H-4': ('1003562', 2296.00),
            'Pole,wood,50 Ft,class H-3': ('1003561', 1869.87),
            'Pole,wood,50 Ft,class H-2': ('1003560', 1177.25),
            'Pole,wood,50 Ft,class H-1': ('1003559', 1193.66),
            'Pole,wood,50 Ft,class 1': ('1616501', 1108.25),
            'Pole,wood,50 Ft,class 2': ('1616502', 1033.43),
            'Pole,wood,50 Ft,class 3': ('1616503', 863.43),
            'Pole,wood,55 Ft,class H-3': ('1003568', 2281.50),
            'Pole,wood,55 Ft,class H-2': ('1003567', 1530.46),
            'Pole,wood,55 Ft,class H-1': ('1003566', 1633.85),
            'Pole,wood,55 Ft,class 1': ('1616551', 1230.81),
            'Pole,wood,55 Ft,class 2': ('1616552', 1107.40),
            'Pole,wood,55 Ft,class 3': ('1616553', 965.07),
            'Pole,wood,60 Ft,class H-4': ('1003576', 3189.00),
            'Pole,wood,60 Ft,class H-3': ('1003575', 2882.52),
            'Pole,wood,60 Ft,class H-2': ('1616609', 2216.16),
            'Pole,wood,60 Ft,class H-1': ('1616608', 1618.26),
            'Pole,wood,60 Ft,class 2': ('1616602', 1252.31),
            'Pole,wood,60 Ft,class 3': ('1616603', 935.77),
            'Pole,wood,65 Ft,class H-3': ('1003579', 2342.57),
            'Pole,wood,70 Ft,class H-4': ('1003584', 3354.59),
            'Pole,wood,70 Ft,class H-1': ('1616708', 2687.30),
            'Pole,wood,75 Ft,class H-1': ('1616758', 2769.95),
            'Pole,wood,75 Ft,class 1': ('1616751', 2630.39),
            'Pole,wood,80 Ft,class H-4': ('1003593', 3655.50),

            # Transformers
            'Single Phase XFMR Tank 2.4 120/240 10kva': ('4500349', 677.45),
            'Single Phase XFMR Tank 2.4 120/240 15kva': ('4500488', 917.92),
            'Single Phase XFMR Tank 2.4 120/240 25kva': ('4500350', 846.68),
            'Single Phase XFMR Tank 7.2 120/240 10kva': ('4500353', 1373.65),
            'Single Phase XFMR Tank 7.2 120/240 15kva': ('4500497', 473.64),
            'Single Phase XFMR Tank 7.2 120/240 25kva': ('4500354', 1719.49),
            'Single Phase XFMR Tank 7.2 120/240 37.5kva': ('4500498', 896.39),
            'Single Phase XFMR Tank 7.2 120/240 50kva': ('4500355', 1887.82),
            'Single Phase XFMR Tank 19.9 120/240 10 kva': ('4500314', 933.17),
            'Single Phase XFMR Tank 19.9 120/240 25 kva': ('4500315', 1162.05),
            'Single Phase XFMR Tank 19.9 120/240 50 kva': ('4500316', 1552.73),
            'Single Phase XFMR Tank 19.9 120/240 75 kva': ('4500317', 2639.31),
            'Single Phase XFMR Tank 19.9 120/240 100 kva': ('4500318', 3323.88),
            'Single Phase XFMR Tank 19.9 120/240 167 kva': ('4500319', 4478.38),

            # Conduit
            'Conduit PVC 2" Sch 80': ('6619001', 5.69),
            'Conduit PVC 3" Sch 80': ('6619209', 10.65),
            'Conduit PVC 4" Sch 80': ('6619308', 24.03),
            'Conduit PVC 6" Sch 80': ('6619506', 16.76),

            # S-Series Poles and Components
            'S-02.0 / C3-30-0': ('1007995', 1243.00),
            'S-02.0 / C2-40-0': ('1008000', 2026.51),
            'S-02.0 / C3-40-0': ('1007999', 335.00),
            'S-02.0 / C3-45-0': ('1008002', 2763.48),
            'S-02.4 / C2-45-0': ('1008003', 2845.75),
            'S-02.9-C1-45-0': ('8006412', 3359.58),
            'S-02.0 / C3-50-0': ('1008005', 2583.19),
            'S-02.4 / C2-50-0': ('1008006', 2694.26),
            'S-02.9-C1-50-0': ('8006413', 3608.12),
            'S-04.2 / H2-050': ('7991537', 4630.23),
            'S-04.2 / H2-060': ('7991549', 6899.00),
            'S-04.9 / H3-050': ('7991538', 2058.63),
            'S-05.7 / H4-050': ('7991617', 2168.95),
            'S-02.0 / C3-55-0': ('1008007', 2661.09),
            'S-02.4 / C2-55-0': ('1008008', 2259.10),
            'S-02.9-C1-55-0': ('1008009', 3720.84),
            'S-03.5-H1-55': ('7991542', 2018.87),
            'S-04.2 / H2-055': ('7991543', 5905.60),
            'S-05.7 / H4-055': ('7991623', 2457.41),
            'S-02.0 / C3-60-0': ('1008010', 2641.06),
            'S-02.4 / C2-60-0': ('1008011', 2421.98),
            'S-02.9 / C1-60-0': ('1008012', 3586.53),
            'S-02.9-H1-50': ('7991536', 4997.00),
            'S-02.9-C1-55': ('7991541', 4273.00),
            'S-02.9-C1-60': ('7991547', 5538.51),
            'S-03.5-H1-60': ('7991548', 1889.54),
            'S-05.7 / H4-060': ('7991629', 2312.65),
            'S-06.5-H5-060': ('7991630', 9259.26),
            'S-02.9-C1-65': ('7991553', 5425.00),
            'S-03.5-H1-65': ('7991554', 8209.38),
            'S-04.2 / H2-065': ('7991555', 7223.00),
            'S-04.9-H3-65': ('7991556', 2421.20),
            'S-07.4-H6-050': ('7991619', 2529.97),
            'S-07.4-H6-055': ('7991625', 2863.91),
            'S-09.0-H8-060': ('7991633', 3509.33),
            'S-05.7-H4-65': ('7991635', 2576.63),
            'S-06.5-H5-055': ('7991624', 8068.32),
            'S-06.5-H5-065': ('7991636', 2954.10),
            'S-02.9-C1-70': ('7991559', 11378.38),
            'S-06.5-H5-050': ('7991618', 2232.99),
            'S-04.9 / H3-055': ('7991544', 2089.18),
            'S-03.5-H1-70': ('7991560', 2356.27),
            'S-04.2-H2-70': ('7991561', 5971.22),
            'S-04.9 / H3-070': ('7991562', 2685.00),
            'S-06.5-H5-70': ('7991642', 3291.25),
            'S-10.0-H9-70': ('7991646', 4350.00),
            'S-02.9-C1-75': ('7991565', 6194.00),
            'S-03.5-H1-75': ('7991566', 5690.00),
            'S-04.2 / H2-075': ('7991567', 2700.00),
            'S-04.9 / H3-075': ('7991568', 6909.39),
            'S-09.0-H8-75': ('7991651', 4540.00),
            'S-10.0-H9-75': ('7991652', 4356.25),
            'S-02.9-C1-80': ('7991571', 6631.00),
            'S-03.5-H1-80': ('7991572', 7873.00),
            'S-04.2 / H2-080': ('7991573', 10711.20),
            'S-04.9 / H3-080': ('7991574', 3223.75),
            'S-05.7 / H4-080': ('7991653', 4133.86),
            'S-08.0-H7-80': ('7991656', 7883.00),
            'S-10.0-H9-80': ('7991658', 11501.00),
            'S-02.9-C1-85': ('7991577', 7174.00),
            'S-04.2 / H2-085': ('7991579', 3213.75),
            'S-06.5-H5-085': ('7991660', 22869.58),
            'S-05.7 / H4-052': ('7991659', 12229.96),
            'S-02.9-C1-90': ('7991583', 8702.00),
            'S-03.5-H1-90': ('7991584', 9505.00),
            'S-04.9-H3-90': ('7991586', 8774.00),
            'S-02.9-C1-95': ('7991588', 8654.65),
            'S-04.2 / H2-095': ('7991590', 10977.00),
            'S-06.5-H5-095': ('7991672', 16413.98),
            'S-02.9-C1-100': ('7991593', 9340.56),
            'S-06.5-H5-105': ('7991684', 18209.52),

            # Special Components and Regulators
            'Regulator, V, 1PH, LINE, 720': ('7999445', 11796.13),
            'Pole, Steel, Self-Supporting Deadend, Direct Buried': ('7991948', 11925.37),

            # Special IDs for Alley Arms
            '8004115': ('8004115', 0),
            '8004116': ('8004116', 0)
        }
        
        # Add framing requirements dictionary
        self.framing_requirements = {
            # 1PH Configurations
            'EH101 - 1PH TAN': {
            'Pin Insulator': 1
            },
            'EH106 - 1PH ANGLE': {
            'Pin Insulator': 1
            },
            'EH111 - 1PH DDE': {
            'Pin Insulator': 1,
            'DE Insulator': 2
            },
            'EH121 - 1PH CORNER': {
            'DE Insulator': 2
            },
            'EH131 - 1PH DE': {
            'DE Insulator': 1
            },
            'EH151 - 1PH DDE FOR FUSING': {
            'DE Insulator': 2
            },
            'EH201 - 1PH DDE FOR LINE EXTENSION': {
            'Pin Insulator': 1,
            'DE Insulator': 2
            },
            'EH221 - 1PH TAP FROM 1PH LINE': {
            'DE Insulator': 1
            },
            'EH226 - 1PH TAP FROM 3PH LINE': {
            'DE Insulator': 1
            },
            'EH231 - 1PH TAP FROM 1 PH LINE FOR FUSING': {
            'DE Insulator': 1
            },
            'EH236 - 1PH TAP FROM 3PH LINE FOR FUSING': {
            'DE Insulator': 1
            },
            'EH301 - 1PH HORIZIONTAL TANGENT': {
            'Pin Insulator': 2,
            '2PO Tan xarm': 1
            },
            'EH321 - 1PH HORIZIONTAL DDE': {
            'Pin Insulator': 2,
            '2PO DE xarm': 1,
            'DE Insulator': 4
            },
            'EH331 - 1PH HORIZIONTAL CORNER': {
            'Pin Insulator': 2,
            '2PO DE xarm': 2,
            'DE Insulator': 4
            },
            'EH341 - 1PH HORIZIONTAL DE': {
            '2PO Tan xarm': 1,
            'DE Insulator': 2
            },
            'EH401 - 1PH HORIZIONTAL DDE LINE EXTENSION': {
            'Pin Insulator': 2,
            '2PO DE xarm': 2,
            'DE Insulator': 4
            },
            'EH421 - 1PH HORIZONTAL TAP FROM 1PH LINE': {
            'Pin Insulator': 3,
            '2PO DE xarm': 1,
            'DE Insulator': 2,
            '2PO Tan xarm': 1
            },

            # 2PH Configurations
            'EI101 - 2PH TAN': {
            'Pin Insulator': 2,
            '2PO DE xarm': 1
            },
            'EI131 - 2PH DDE': {
            'Pin Insulator': 2,
            'DE Insulator': 4,
            '2PO DE xarm': 1
            },
            'EI141 - 2PH CORNER': {
            'Pin Insulator': 2,
            'DE Insulator': 4,
            '2PO DE xarm': 2
            },
            'EI151 - 2PH DE': {
            'DE Insulator': 2,
            '2PO DE xarm': 1
            },
            'EI171 - 2PH DDE FOR FUSING': {
            'DE Insulator': 4,
            '2PO DE xarm': 1
            },
            'EI201 - 2PH TAP ONE DIRECTION': {
            'Pin Insulator': 1,
            'DE Insulator': 2,
            '2PO DE xarm': 1
            },
            'EI221 - 2PH TAP TWO DIRECTIONS': {
            'Pin Insulator': 3,
            '2PO DE xarm': 1
            },
            'EI231 - 2PH LINE EXTENSION': {
            'Pin Insulator': 2,
            'DE Insulator': 2,
            '4PO DE xarm': 1
            },
            'EI401 - 2PH TAN HIGH NEUTRAL': {
            'Pin Insulator': 2,
            '4PO DE xarm': 1
            },
            'EI431 - 2PH DDE HIGH NEUTRAL': {
            'Pin Insulator': 2,
            'DE Insulator': 4,
            '4PO DE xarm': 1
            },

            # 3PH Configurations
            'EJ300 TAN': {
            'Pin Insulator': 4,
            '4PO DE xarm': 1
            },
            'EJ300 ANGLE': {
            'Pin Insulator': 4,
            '4PO DE xarm': 1
            },
            'EJ300 DE': {
            'DE Insulator': 5,
            '4PO DE xarm': 1
            },
            'EJ300 DDE': {
            'Pin Insulator': 4,
            'DE Insulator': 10,
            '4PO DE xarm': 1
            },
            'EJ300 CORNER': {
            'Pin Insulator': 4,
            'DE Insulator': 9,
            '4PO DE xarm': 2
            },
            'EJ376 - 3PH TAP FROM 3PH LINE': {
            'Pin Insulator': 2,
            'DE Insulator': 3,
            '4PO DE xarm': 1
            },
            'EJ371 - 3PH DDE LINE EXTENSION': {
            'Pin Insulator': 4,
            'DE Insulator': 10,
            '4PO DE xarm': 1
            },
            'EJ601 - 3PH TAN ALLEY ARM': {
            'Pin Insulator': 3,
            '8004115': 1
            },
            'EJ606 - 3PH TAN ALLEY ARM': {
            'Pin Insulator': 3,
            '8004116': 1
            },
            'EJ621 - 3PH TAN ALLEY ARM HIGH NEUTRAL': {
            'Pin Insulator': 3,
            '8004115': 1,
            'DE Insulator': 2
            },
            'EJ800 TAN': {
            'Pin Insulator': 6,
            '4PO DE xarm': 2
            },
            'EJ800 ANGLE': {
            'Pin Insulator': 6,
            '4PO DE xarm': 2
            },
            'EJ800 DE': {
            'DE Insulator': 8,
            '4PO DE xarm': 2
            },
            'EJ800 DDE': {
            'Pin Insulator': 6,
            'DE Insulator': 12,
            '4PO DE xarm': 2
            },
            'EJ800 CORNER': {
            'Pin Insulator': 6,
            'DE Insulator': 12,
            '4PO DE xarm': 3
            },
            'EJ861 - 3PH TAP FROM 3PH LINE': {
            'Pin Insulator': 3,
            'DE Insulator': 4,
            '4PO DE xarm': 2
            },
            'EJ862 - 3PH TWO DIRECTIONAL TAP FROM 3PH LINE': {
            'Pin Insulator': 4,
            'DE Insulator': 6,
            '4PO DE xarm': 2
            },
            'EJ871 - 3PH DDE LINE EXTENSION - RAPTOR AREA': {
            'Pin Insulator': 6,
            'DE Insulator': 12,
            '4PO DE xarm': 2,
            'Connector cover (large)': 6,
            'Deadend cover': 12
            },
            'EJ909 - 3PH TAN UNDERBUILD': {
            'Pin Insulator': 6,
            '4PO DE xarm': 2,
            '2PO Tan xarm': 1
            }
        }

    def clean_framing_string(self, framing_str):
        """Clean framing strings by removing class types and standardizing format"""
        if not framing_str or not isinstance(framing_str, str):
            return ""
            
        # Remove ST50, H1, class types, etc.
        parts = framing_str.split()
        framing = parts[0]  # Get first part
        
        # Add TAN if it exists in the original string
        if 'TAN' in framing_str:
            framing += ' TAN'
            
        return framing

    def clean_pole_type(self, pole_str):
        """Clean pole type strings to match database format"""
        if not pole_str or not isinstance(pole_str, str):
            return ""
            
        # Handle DIST-S prefix
        if pole_str.startswith('DIST-S-'):
            pole_str = pole_str[7:]  # Remove 'DIST-S-'
            
        return pole_str

    def safe_set_cell_value(self, sheet, cell_ref, value):
        """Safely set cell value handling merged cells"""
        merged_ranges = sheet.merged_cells.ranges
        target_cell = sheet[cell_ref]
        
        for merged_range in merged_ranges:
            if target_cell.coordinate in merged_range:
                top_left = merged_range.start_cell
                top_left.value = value
                return
        
        target_cell.value = value

    def apply_safety_factor(self, quantity, item_type):
        """Apply appropriate safety factors based on item type"""
        if 'ACSR' in item_type or 'AAC' in item_type:
            return round(quantity * 1.05)  # 5% safety factor for conductors
        elif 'guy wire' in item_type.lower():
            total_length = quantity * 50  # 50 ft per guy
            return int((total_length / 200) + 1)  # Safety factor for guy wire spools
        elif 'Conduit' in item_type:
            return int(quantity / 10) + (1 if quantity % 10 > 0 else 0)  # 10' sections
        return quantity

    def process_poles(self, data_sheet):
        """Process pole data from column N"""
        pole_counts = {}
        
        for row in data_sheet.iter_rows(min_row=2, values_only=True):
            pole_type = row[13]  # Column N
            if pole_type and isinstance(pole_type, str):
                cleaned_pole = self.clean_pole_type(pole_type)
                
                if cleaned_pole not in pole_counts:
                    # Try to find matching pole in material_data
                    for db_pole, (stock_num, cost) in self.material_data.items():
                        if cleaned_pole in db_pole:
                            pole_counts[cleaned_pole] = {
                                'count': 1,
                                'stock_num': stock_num,
                                'unit_cost': cost
                            }
                            break
                else:
                    pole_counts[cleaned_pole]['count'] += 1
        
        return pole_counts

    def process_conductors(self, data_sheet):
        """Process conductor data with safety factors"""
        conductor_data = {}
        
        for row in data_sheet.iter_rows(min_row=2, values_only=True):
            if row[5] and isinstance(row[5], str):  # Column F
                conductor_type = row[5]
                if any(x in conductor_type for x in ['ACSR', 'AAC']):
                    length = float(row[2]) if row[2] else 0  # Column C
                    if conductor_type not in conductor_data:
                        stock_num, unit_cost = self.material_data.get(conductor_type, ('', 0))
                        conductor_data[conductor_type] = {
                            'length': length,
                            'stock_num': stock_num,
                            'unit_cost': unit_cost
                        }
                    else:
                        conductor_data[conductor_type]['length'] += length
        
        # Apply safety factor
        for conductor in conductor_data.values():
            conductor['length'] = self.apply_safety_factor(conductor['length'], 'ACSR')
        
        return conductor_data

    def process_framing_and_materials(self, data_sheet):
        """Process framing and calculate required materials"""
        framing_counts = {}
        material_counts = {}
        
        for row in data_sheet.iter_rows(min_row=2, values_only=True):
            primary_framing = row[9]   # Column J
            secondary_framing = row[10] # Column K
            
            # Process both primary and secondary framing
            for framing_str in [primary_framing, secondary_framing]:
                if framing_str and isinstance(framing_str, str):
                    cleaned_framing = self.clean_framing_string(framing_str)
                    
                    if cleaned_framing in self.framing_requirements:
                        # Count framing
                        if cleaned_framing not in framing_counts:
                            stock_num, unit_cost = self.material_data.get(cleaned_framing, ('', 0))
                            framing_counts[cleaned_framing] = {
                                'count': 1,
                                'stock_num': stock_num,
                                'unit_cost': unit_cost
                            }
                        else:
                            framing_counts[cleaned_framing]['count'] += 1
                        
                        # Process associated materials
                        required_materials = self.framing_requirements[cleaned_framing]
                        for material, quantity in required_materials.items():
                            if material not in material_counts:
                                stock_num, unit_cost = self.material_data.get(material, ('', 0))
                                material_counts[material] = {
                                    'count': quantity,
                                    'stock_num': stock_num,
                                    'unit_cost': unit_cost
                                }
                            else:
                                material_counts[material]['count'] += quantity
        
        return framing_counts, material_counts

    def generate_sheet(self):
        """Generate the sheet using the template"""
        try:
            # Load template
            template_wb = load_workbook(self.template_path)
            sheet = template_wb.active
            
            # Fill in basic information
            self.safe_set_cell_value(sheet, 'C12', self.wo_number)
            self.safe_set_cell_value(sheet, 'E12', "5R234 Plumtree Ln Pt 2")

            # Process source data
            data_sheet = self.source_workbook.active
            current_row = 16  # Start after headers
            
            # Create a list to store all items
            all_items = []
            
            # Process poles
            pole_data = self.process_poles(data_sheet)
            for pole_type, info in pole_data.items():
                if info['stock_num']:  # Only add if we have valid stock number
                    all_items.append({
                        'type': 'pole',
                        'description': pole_type,
                        'stock_num': info['stock_num'],
                        'quantity': info['count'],
                        'unit': 'EA',
                        'unit_cost': info['unit_cost']
                    })

            # Process conductors
            conductor_data = self.process_conductors(data_sheet)
            for cond_type, info in conductor_data.items():
                all_items.append({
                    'type': 'conductor',
                    'description': cond_type,
                    'stock_num': info['stock_num'],
                    'quantity': info['length'],
                    'unit': 'FT',
                    'unit_cost': info['unit_cost']
                })

            # Process framing and materials
            framing_data, material_data = self.process_framing_and_materials(data_sheet)
            
            # Add framing items
            for frame_type, info in framing_data.items():
                if info['stock_num']:
                    all_items.append({
                        'type': 'framing',
                        'description': frame_type,
                        'stock_num': info['stock_num'],
                        'quantity': info['count'],
                        'unit': 'EA',
                        'unit_cost': info['unit_cost']
                    })
            
            # Add material items
            for material_type, info in material_data.items():
                if info['stock_num']:
                    all_items.append({
                        'type': 'material',
                        'description': material_type,
                        'stock_num': info['stock_num'],
                        'quantity': info['count'],
                        'unit': 'EA',
                        'unit_cost': info['unit_cost']
                    })

            # Sort and write items
            for idx, item in enumerate(all_items, 1):
                row = current_row + idx - 1
                sheet.cell(row=row, column=1).value = idx  # Item number
                sheet.cell(row=row, column=2).value = item['stock_num']
                sheet.cell(row=row, column=3).value = item['quantity']
                sheet.cell(row=row, column=4).value = item['unit']
                sheet.cell(row=row, column=5).value = item['description']
                sheet.cell(row=row, column=10).value = item['unit_cost']
                sheet.cell(row=row, column=11).value = f'=C{row}*J{row}'

            # Save the workbook
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Long_Lead_Items_{self.wo_number}.xlsx"
            )
            if file_path:
                template_wb.save(file_path)
                messagebox.showinfo("Success", f"File saved successfully as {file_path}")
                return file_path

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            raise