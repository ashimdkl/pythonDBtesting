import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from ttkthemes import ThemedTk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import math
import mergeXML
import shutil
import os
import customtkinter as ctk
from mergeXML import XMLTagExtractorApp  # Adjust import path as needed


class DataExtractionApp:
    def __init__(self, root):
        # Initialize the application window and set up basic variables and GUI components
        self.root = root
        self.root.title("Data Extraction App")
        self.root.geometry('1100x650')

        # Variables to hold file paths, columns, and selected data
        self.file_path = None
        self.columns = []
        self.selected_columns = []
        self.step = 1  # Step in the process
        self.output_data = []
        self.max_force_data = {}  # Dictionary to hold maximum force data by sequence number
        self.soil_class_data = {}  # New dictionary to store soil class data

        # Set up the graphical user interface (GUI)
        self.setup_gui()

    def setup_gui(self):
        # Import customtkinter if used
        try:
            use_ctk = True
            # Set the appearance mode and default color theme
            ctk.set_appearance_mode("light")
            ctk.set_default_color_theme("blue")
        except ImportError:
            use_ctk = False
    
        # Configure basic ttk style for fallback
        style = ttk.Style()
        style.configure('TFrame', background='#ffffff')
        style.configure('TLabel', background='#ffffff', font=('Helvetica', 12), foreground='#333333')
        style.configure('TButton', font=('Helvetica', 10), borderwidth=1)
        style.map('TButton', background=[('active', '#e1e1e1')], foreground=[('active', '#000000')])

        # Create header frame
        header_frame = ttk.Frame(self.root, style='TFrame')
        header_frame.pack(fill="x", padx=20, pady=(20, 0))
        
        # App title and logo
        app_title = ttk.Label(
            header_frame, 
            text="Part 1: Data Extraction Tool",
            font=("Helvetica", 20, "bold"),
            foreground="#0066cc"
        )
        app_title.pack(side="left")
        
        # Main container with two panels
        container = ttk.Frame(self.root, style='TFrame')
        container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Left sidebar for step navigation
        self.sidebar = ttk.Frame(container, style='TFrame', width=250)
        self.sidebar.pack(side="left", fill="y", padx=(0, 20))
        self.sidebar.pack_propagate(False)  # Maintain width
        
        # Step indicators in sidebar
        steps_label = ttk.Label(
            self.sidebar,
            text="Process Steps",
            font=("Helvetica", 14, "bold"),
            foreground="#444444"
        )
        steps_label.pack(anchor="w", pady=(0, 15))
        
        # Create step buttons
        self.step_buttons = []
        steps = [
            "1. Upload HIS Excel",
            "2. Fusing Coordination",
            "3. XML File Parsing",
            "4. Soil Class Data"
        ]
        
        for i, step_text in enumerate(steps, 1):
            step_frame = ttk.Frame(self.sidebar, style='TFrame')
            step_frame.pack(fill="x", pady=5)
            
            # Circle indicator
            indicator = tk.Canvas(step_frame, width=30, height=30, bg="#ffffff", highlightthickness=0)
            indicator.pack(side="left", padx=(0, 10))
            
            # Draw circle
            if i == self.step:
                # Active step
                indicator.create_oval(5, 5, 25, 25, fill="#0066cc", outline="#0066cc")
                indicator.create_text(15, 15, text=str(i), fill="white", font=("Helvetica", 10, "bold"))
                text_color = "#0066cc" 
            else:
                # Inactive step
                indicator.create_oval(5, 5, 25, 25, fill="#ffffff", outline="#aaaaaa")
                indicator.create_text(15, 15, text=str(i), fill="#aaaaaa", font=("Helvetica", 10))
                text_color = "#666666"
            
            # Step text
            step_label = ttk.Label(
                step_frame,
                text=step_text,
                font=("Helvetica", 12),
                foreground=text_color
            )
            step_label.pack(side="left", anchor="w")
            
            self.step_buttons.append((indicator, step_label))
        
        # Main content area
        self.main_frame = ttk.Frame(container, style='TFrame')
        self.main_frame.pack(side="left", fill="both", expand=True)
        
        # Step indicator in the main frame
        self.step_label = ttk.Label(
            self.main_frame,
            text="Step 1: Upload your Hendrix Input Sheet (HIS)",
            font=("Helvetica", 16, "bold"),
            foreground="#333333"
        )
        self.step_label.pack(anchor="w", pady=(0, 20))
        
        # Step description
        step_description = ttk.Label(
            self.main_frame,
            text="Select an Excel file containing the Hendrix Input Sheet data.\nThe required columns will be automatically selected.",
            font=("Helvetica", 12),
            wraplength=600
        )
        step_description.pack(anchor="w", pady=(0, 20))
        
        # Upload button with modern styling
        upload_frame = ttk.Frame(self.main_frame, style='TFrame')
        upload_frame.pack(anchor="w", pady=(0, 20))
        
        if use_ctk:
            # If customtkinter is available, use it for buttons
            self.upload_btn = ctk.CTkButton(
                upload_frame,
                text="📄 Upload HIS Excel File",
                command=self.upload_file,
                height=40,
                corner_radius=8,
                font=("Helvetica", 12),
                fg_color="#0066cc",
                hover_color="#004d99"
            )
        else:
            # Fallback to ttk
            self.upload_btn = ttk.Button(
                upload_frame,
                text="📄 Upload HIS Excel File",
                command=self.upload_file,
                style='TButton'
            )
        self.upload_btn.pack(side="left")
        
        # Column selection label
        self.column_label = ttk.Label(
            self.main_frame,
            text="Selected Columns:",
            font=("Helvetica", 12, "bold")
        )
        self.column_label.pack(anchor="w", pady=(20, 5))
        
        # Listbox frame with shadow effect
        self.listbox_container = ttk.Frame(self.main_frame, style='TFrame')
        self.listbox_container.pack(fill="both", expand=True, pady=(0, 20))
        
        self.listbox_frame = ttk.Frame(self.listbox_container, style='TFrame')
        self.listbox_frame.pack(fill="both", expand=True, padx=2, pady=2)
        
        # Create listbox with custom styling
        self.column_listbox = tk.Listbox(
            self.listbox_frame,
            selectmode=tk.MULTIPLE,
            font=("Helvetica", 11),
            borderwidth=1,
            relief="solid",
            selectbackground="#0066cc",
            highlightthickness=1,
            highlightcolor="#cccccc"
        )
        self.column_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Modern scrollbar
        self.scrollbar = ttk.Scrollbar(
            self.listbox_frame,
            orient=tk.VERTICAL,
            command=self.column_listbox.yview
        )
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.column_listbox.config(yscrollcommand=self.scrollbar.set)
        
        # Text area for pasting data
        self.paste_label = ttk.Label(
            self.main_frame,
            text="Paste Data Here:",
            font=("Helvetica", 12, "bold")
        )
        
        self.text_frame = ttk.Frame(self.main_frame, style='TFrame')
        
        self.paste_text = tk.Text(
            self.text_frame,
            wrap=tk.WORD,
            height=15,
            font=("Helvetica", 11),
            borderwidth=1,
            relief="solid",
            padx=10,
            pady=10,
            bg="#f9f9f9"
        )
        self.paste_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.text_scrollbar = ttk.Scrollbar(
            self.text_frame,
            orient=tk.VERTICAL,
            command=self.paste_text.yview
        )
        self.text_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.paste_text.config(yscrollcommand=self.text_scrollbar.set)
        
        # Initially hide paste section
        self.paste_label.pack_forget()
        self.text_frame.pack_forget()
        
        # Button container at the bottom
        self.button_frame = ttk.Frame(self.main_frame, style='TFrame')
        self.button_frame.pack(fill="x", pady=(20, 0))
        
        # Action buttons with modern styling
        if use_ctk:
            self.process_btn = ctk.CTkButton(
                self.button_frame,
                text="Process Data and Continue",
                command=self.parse_and_next_step,
                height=40,
                corner_radius=8,
                font=("Helvetica", 12),
                fg_color="#0066cc",
                hover_color="#004d99"
            )
            
            self.skip_btn = ctk.CTkButton(
                self.button_frame,
                text="Skip This Step",
                command=self.next_step,
                height=40,
                corner_radius=8,
                font=("Helvetica", 12),
                fg_color="#f0f0f0",
                text_color="#444444",
                hover_color="#e0e0e0",
                border_width=1,
                border_color="#dddddd"
            )
            
            self.next_btn = ctk.CTkButton(
                self.button_frame,
                text="Generate Final Report",
                command=self.generate_report,
                height=48,
                corner_radius=8,
                font=("Helvetica", 14, "bold"),
                fg_color="#28a745",
                hover_color="#218838"
            )
        else:
            self.process_btn = ttk.Button(
                self.button_frame,
                text="Process Data and Continue",
                command=self.parse_and_next_step,
                style='TButton'
            )
            
            self.skip_btn = ttk.Button(
                self.button_frame,
                text="Skip This Step",
                command=self.next_step,
                style='TButton'
            )
            
            self.next_btn = ttk.Button(
                self.button_frame,
                text="Generate Final Report",
                command=self.generate_report,
                style='TButton'
            )
        
        self.process_btn.pack(side="left", padx=(0, 10))
        self.skip_btn.pack(side="left")
        
        # Hide the generate report button initially
        self.next_btn.pack_forget()
        
        # Status frame for feedback
        status_frame = ttk.Frame(self.root, style='TFrame')
        status_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        # Status label for user feedback
        self.status_label = ttk.Label(
            status_frame,
            text="Ready to start processing",
            font=("Helvetica", 10),
            foreground="#666666"
        )
        self.status_label.pack(anchor="w")
        
        # Output frame
        self.output_frame = ttk.Frame(self.root, style='TFrame')
        self.output_frame.pack(expand=True, fill="both")

    def start_analysis(self):
        # Transition from the intro frame to the main frame where analysis starts
        self.intro_frame.pack_forget()
        self.main_frame.pack(expand=True, fill="both")

    def upload_file(self):
        # Open file dialog to select file appropriate for current step
        if self.step == 1:
            filetypes = [("Excel files", "*.xlsx *.xls")]
        else:
            filetypes = [("XML files", "*.xml"), ("Text files", "*.txt")]
            
        self.file_path = filedialog.askopenfilename(filetypes=filetypes)
        
        if not self.file_path:
            return
            
        # Call the appropriate function based on the step of the process
        if self.step == 1:
            self.load_columns_from_file()
            # Updated to handle step 1 file uploads
            self.parse_data()
            self.next_step()
            self.status_label.configure(text=f"Processed Excel file: {os.path.basename(self.file_path)}")
        elif self.step == 3:
            # For step 3, we process the XML file differently
            self.process_large_xml(self.file_path)
        elif self.step == 4:
            # For step 4's joint support XML
            self.parse_step7_joint_support()
            self.status_label.configure(text="Joint support data processed successfully")

    def process_large_xml(self, file_path):
        try:
            # Create a Toplevel window instead of passing None
            xml_window = tk.Toplevel(self.root)
            xml_processor = XMLTagExtractorApp(xml_window)
            
            # Configure the processor window
            xml_window.title("XML Tag Extractor")
            xml_window.geometry("600x400")
            
            # Process the XML
            xml_processor.extract_tags(file_path)
            
            # Close the XML processor window
            xml_window.destroy()
            
            # Verify files exist
            required_files = [
                "XMLextractConstrucStakingReport_framing_type_direction_length.txt",
                "XMLextractPoleType.txt",
                "XMLextractGuyUsage_seq_elementType_usage.txt"
            ]
            
            missing_files = []
            for file_name in required_files:
                if not os.path.exists(file_name):
                    missing_files.append(file_name)
            
            if missing_files:
                raise FileNotFoundError(f"Missing required files: {', '.join(missing_files)}")
            
            messagebox.showinfo("Success", "XML processing complete. Files generated.")
            self.next_step()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process XML file: {e}")

    def load_columns_from_file(self):
        # Load columns from the first row of the Excel file to allow user selection
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active
            self.columns = [cell.value for cell in sheet[1]]
            self.column_listbox.delete(0, tk.END)
            
            # Add all columns to the listbox
            for column in self.columns:
                self.column_listbox.insert(tk.END, column)
            
            # Auto-select the five specific columns we always need
            columns_to_select = ["Sequence", "Facility ID", "Existing Transformers", "Primary Riser", "Secondary Riser"]
            
            # Find and select these columns in the listbox
            for i, column in enumerate(self.columns):
                if column in columns_to_select:
                    self.column_listbox.selection_set(i)
                    
            # Optionally, you can show a message to let the user know columns were auto-selected
            messagebox.showinfo("Columns Selected", "The required columns have been automatically selected.")
                    
        except Exception as e:
            # Show an error if something goes wrong while reading the file
            messagebox.showerror("Error", f"Failed to read file: {e}")

    def parse_data(self):
        # Parse the Excel data based on the columns selected by the user
        if not self.file_path:
            messagebox.showerror("Error", "Please upload a file first!")
            return

        # Get the columns selected by the user
        selected_indices = self.column_listbox.curselection()
        self.selected_columns = [self.column_listbox.get(i) for i in selected_indices]
        if not self.selected_columns:
            messagebox.showerror("Error", "Please select at least one column to keep!")
            return

        try:
            if self.step == 1:
                workbook = load_workbook(self.file_path)
                sheet = workbook.active

                # Open a file to write the selected data
                with open("extractHIS_seq_facID_existingTrans_primaryRiser_secondaryRiser.txt", "w") as file:
                    file.write("\t".join(self.selected_columns) + "\n")
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0] is not None:  # Assuming sequence number is the first column
                            row_data = [str(row[self.columns.index(col)]) for col in self.selected_columns]
                            file.write("\t".join(row_data) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse file: {e}")

    def next_step(self):
        # Move to the next step in the process
        self.step += 1
        self.update_step_buttons()  # Update step indicators
        self.file_path = None
        self.columns = []
        self.selected_columns = []
        self.column_listbox.delete(0, tk.END)

        # Clear previous content
        self.paste_label.pack_forget()
        self.text_frame.pack_forget() 
        self.upload_btn.pack_forget()
        self.column_label.pack_forget()
        self.listbox_container.pack_forget()
        self.next_btn.pack_forget()

        # Adjust GUI elements based on the current step
        if self.step == 2:
            self.step_label.config(text="Step 2: Copy and Paste your Fusing Coordination Report")
            self.paste_label.config(text="Paste Fusing Coordination Data Here:")
            self.paste_label.pack(anchor="w", pady=(20, 5))
            self.text_frame.pack(fill="both", expand=True, pady=(0, 20))
            self.process_btn.configure(text="Process Data and Continue")
            
        elif self.step == 3:
            self.step_label.config(text="Step 3: Upload your Large XML File for Parsing")
            self.upload_btn.configure(text="Upload Large XML File")
            self.upload_btn.pack(anchor="w", pady=(0, 20))
            self.process_btn.configure(text="Parse Data and Continue")
            
        elif self.step == 4:
            self.step_label.config(text="Step 4: Soil Class Data & Joint Support")
            
            # Show paste area for soil class data
            self.paste_label.config(text="Paste Soil Class Data Here:")
            self.paste_label.pack(anchor="w", pady=(20, 5))
            self.text_frame.pack(fill="both", expand=True, pady=(0, 20))
            
            # Show upload button for joint support XML
            self.upload_btn.configure(text="Upload Joint Support XML")
            self.upload_btn.pack(anchor="w", pady=(20, 10))
            
            # Change process button to parse soil data
            self.process_btn.configure(text="Parse Soil Class Data", command=self.parse_soil_class_data)
            
            # Show generate report button 
            self.next_btn.pack(pady=(20, 0))



    def update_step_buttons(self):
        """Update the step indicators when changing steps"""
        for i, (indicator, label) in enumerate(self.step_buttons, 1):
            indicator.delete("all")  # Clear existing canvas
            
            if i == self.step:
                # Active step
                indicator.create_oval(5, 5, 25, 25, fill="#0066cc", outline="#0066cc")
                indicator.create_text(15, 15, text=str(i), fill="white", font=("Helvetica", 10, "bold"))
                label.configure(foreground="#0066cc")
            else:
                # Inactive step
                indicator.create_oval(5, 5, 25, 25, fill="#ffffff", outline="#aaaaaa")
                indicator.create_text(15, 15, text=str(i), fill="#aaaaaa", font=("Helvetica", 10))
                label.configure(foreground="#666666")

    def parse_and_next_step(self):
        # Parse the data and move to the next step based on the current step
        if self.step == 1:
            self.parse_data()
        elif self.step == 2:
            self.parse_pasted_data() 
        elif self.step == 3:
            self.process_large_xml(self.file_path)
            return  # Don't call next_step() here as process_large_xml will do it
        elif self.step == 4:
            self.parse_step7_joint_support()
        
        # Move to next step after processing  
        self.next_step()

    def parse_pasted_data(self):
        # Parse data pasted by the user into the text box
        pasted_data = self.paste_text.get("1.0", tk.END).strip()
        if not pasted_data:
            messagebox.showerror("Error", "Please paste data into the text box.")
            return

        try:
            # Split the pasted data into lines and extract relevant fields using regular expressions
            lines = pasted_data.split("\n")
            header = lines[0].split("\t")
            data = []
            for line in lines[1:]:
                fields = line.split("\t")
                sequence = fields[0][:4]  # Assuming sequence is the first 4 characters
                existing = fields[2]  # Assuming existing value is the second field
                data.append([sequence, existing])

            # Save the extracted data to a file
            with open("extractFusingCoordination_newOrExistingFusing.txt", "w") as file:
                file.write("Sequence\tExisting\n")
                for row in data:
                    file.write("\t".join(row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse pasted data: {e}")

    def parse_soil_class_data(self):
        # Parse soil class data pasted by the user into the text box
        pasted_data = self.paste_text.get("1.0", tk.END).strip()
        if not pasted_data:
            messagebox.showerror("Error", "Please paste data into the text box.")
            return

        try:
            # Regular expression pattern to extract soil class information from the pasted data
            lines = pasted_data.split('\n')
            pattern = r'(\d+)(?:-(\d+))?\s+(\d+)\s+(.*)'
            
            for line in lines:
                match = re.match(pattern, line.strip())
                if match:
                    start_seq, end_seq, soil_class, description = match.groups()
                    start_seq = int(start_seq)
                    end_seq = int(end_seq) if end_seq else start_seq

                    # Store the soil class information for each sequence in the range
                    for seq in range(start_seq, end_seq + 1):
                        self.soil_class_data[seq] = {'soil_class': soil_class, 'description': description}

            messagebox.showinfo("Success", "Soil class data parsed successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse soil class data: {e}")

        # Now, update the MAX force file with soil class information
        self.update_max_force_file_with_soil_class()

    def update_max_force_file_with_soil_class(self):
        # Update the MAX force file with the soil class data parsed earlier
        try:
            with open("extractMAX_sequence_MaxForce.txt", "r") as file:
                lines = file.readlines()

            updated_lines = [lines[0], "Sequence | Max Force | Soil Class\n", "-" * 40 + "\n"]

            for line in lines[2:]:
                parts = line.strip().split('|')
                sequence = int(parts[0].strip())
                max_force = parts[1].strip()
                soil_class = self.soil_class_data.get(sequence, {}).get('soil_class', 'N/A')
                updated_lines.append(f"{sequence:4d} | {max_force:8} | {soil_class}\n")

            # Write the updated lines back to the file
            with open("extractMAX_sequence_MaxForce.txt", "w") as file:
                file.writelines(updated_lines)

            messagebox.showinfo("Success", "MAX force file updated with soil class information.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update MAX force file: {e}")

    def parse_step3_xml(self):
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()
            data = {}
            pole_types = {}

            # Extract data from the XML structure
            for report in root.findall('.//construction_staking_report'):
                sequence = report.find('structure_number').text or ''
                framing = report.find('structure_name').text or ''
                latitude = report.find('latitude').text or ''
                longitude = report.find('longitude').text or ''
                x_easting = report.find('x_easting').text or ''
                y_northing = report.find('y_northing').text or ''
                stake_description = report.find('stake_description').text or ''

                # Track P1 pole types
                if "P1" in stake_description:
                    pole_type = report.find('pole_property_label').text or ''
                    pole_types[sequence] = pole_type

                # Simplify framing
                framing_parts = framing.split(" ", 2)
                if len(framing_parts) > 2:
                    framing = framing_parts[-1]
                    framing = " ".join(framing.split()[:-1])

                if sequence not in data:
                    data[sequence] = []

                # Always log all points, including standalone P1
                data[sequence].append({
                    'framing': framing,
                    'latitude': latitude,
                    'longitude': longitude,
                    'x_easting': x_easting,
                    'y_northing': y_northing,
                    'stake_description': stake_description
                })

            anchor_data = []
            guy_types = ["P2", "PG", "SE", "NG", "CM", "FG"]
            for sequence, points in data.items():
                p1_point = None
                for point in points:
                    if "P1" in point['stake_description']:
                        p1_point = point
                        break

                # Process anchor points if P1 exists
                if p1_point:
                    x_origin = float(p1_point['x_easting'])
                    y_origin = float(p1_point['y_northing'])
                    stake_description_set = set()

                    for point in points:
                        for guy_type in guy_types:
                            if guy_type in point['stake_description']:
                                x_next = float(point['x_easting'])
                                y_next = float(point['y_northing'])
                                lead_length = math.sqrt((x_next - x_origin) ** 2 + (y_next - y_origin) ** 2)
                                theta = math.degrees(math.atan2(y_next - y_origin, x_next - x_origin))
                                direction = self.get_cardinal_direction(theta)
                                descriptions = point['stake_description'].split(',')
                                for description in descriptions:
                                    if description.strip() not in stake_description_set:
                                        stake_description_set.add(description.strip())
                                        anchor_data.append({
                                            'sequence': sequence,
                                            'type': f"P1 to {description.strip()}",
                                            'latitude': point['latitude'],
                                            'longitude': point['longitude'],
                                            'framing': point['framing'],
                                            'anchor_direction': direction,
                                            'lead_length': lead_length
                                        })

                    # Log standalone P1 if no anchors are found
                    if not stake_description_set:
                        anchor_data.append({
                            'sequence': sequence,
                            'type': "P1",
                            'latitude': p1_point['latitude'],
                            'longitude': p1_point['longitude'],
                            'framing': p1_point['framing'],
                            'anchor_direction': '',
                            'lead_length': 0.0
                        })

            # Sort and save anchor data
            anchor_data.sort(key=lambda x: x['sequence'])
            with open("extractConstrucStakingReport_framing_type_direction_length.txt", "w") as file:
                max_lengths = {
                    'sequence': max(len(item['sequence']) for item in anchor_data),
                    'type': max(len(item['type']) for item in anchor_data),
                    'latitude': max(len(str(item['latitude'])) for item in anchor_data),
                    'longitude': max(len(str(item['longitude'])) for item in anchor_data),
                    'framing': max(len(item['framing']) for item in anchor_data),
                    'anchor_direction': max(len(item['anchor_direction']) for item in anchor_data),
                    'lead_length': max(len(f"{item['lead_length']:.2f}") for item in anchor_data)
                }
                headers = [
                    ("Sequence", max_lengths['sequence']),
                    ("Type", max_lengths['type']),
                    ("Latitude", max_lengths['latitude']),
                    ("Longitude", max_lengths['longitude']),
                    ("Framing", max_lengths['framing']),
                    ("Anchor Direction", max_lengths['anchor_direction']),
                    ("Lead Length", max_lengths['lead_length'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for item in anchor_data:
                    row = [
                        f"{item['sequence']:<{max_lengths['sequence']}}",
                        f"{item['type']:<{max_lengths['type']}}",
                        f"{item['latitude']:<{max_lengths['latitude']}}",
                        f"{item['longitude']:<{max_lengths['longitude']}}",
                        f"{item['framing']:<{max_lengths['framing']}}",
                        f"{item['anchor_direction']:<{max_lengths['anchor_direction']}}",
                        f"{item['lead_length']:<{max_lengths['lead_length']}.2f}"
                    ]
                    file.write(" | ".join(row) + "\n")

            # Save pole type data
            with open("extractPoleType.txt", "w") as file:
                file.write("Sequence\tPole Type\n")
                for sequence, pole_type in sorted(pole_types.items()):
                    file.write(f"{sequence}\t{pole_type}\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse XML file: {e}")

    def get_cardinal_direction(self, angle):
        # Convert an angle into a cardinal direction (e.g., N, NE, E)
        if -2.0 < angle <= 2.0:
            return 'E'
        elif 2.0 < angle <= 88.0:
            return 'NE'
        elif 88.0 < angle <= 92.0:
            return 'N'
        elif 92.0 < angle <= 178.0:
            return 'NW'
        elif -88.0 < angle <= -2.0:
            return 'SE'
        elif -92.0 < angle <= -88.0:
            return 'S'
        elif -178.0 < angle <= -92.0:
            return 'SW'
        else:
            return 'W'

    def parse_and_continue_stringing_chart(self):
        # Parse stringing chart data pasted by the user and continue processing with the XML file
        pasted_data = self.paste_text.get("1.0", tk.END).strip()
        if not pasted_data:
            messagebox.showerror("Error", "Please paste data into the text box.")
            return

        try:
            # Use regular expressions to extract relevant sections from the pasted data
            sections = re.findall(r"Stringing Chart Report\n\nCircuit '(.*?)' Section #(.*?) from structure #(.*?) to structure #(.*?),.*?Span\n(.*?)\n\n", pasted_data, re.DOTALL)
            self.output_data = []

            for section in sections:
                circuit_type, section_num, start_seq, end_seq, spans_data = section
                if "Span Guy" in circuit_type:
                    continue  # Skip Span Guy entries from the pasted data

                spans = re.findall(r"\n\s+(\d+\.\d+)\s+", spans_data)
                total_span_length = sum(map(float, spans))
                sequences = f"{start_seq} - {end_seq}"
                self.output_data.append((section_num, sequences, total_span_length, circuit_type))

            messagebox.showinfo("Success", f"Pasted data parsed successfully. Please upload the Span Guy XML file.")
            self.upload_file()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse stringing chart data: {e}")

    def parse_span_guy_xml(self):
        # Parse the XML file for span guy data (wires supporting poles)
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()

            for section in root.findall('.//section_sagging_data'):
                circuit_type = section.find('circuit').text.strip()
                section_num = section.find('sec_no').text.strip()
                start_seq = section.find('from_str').text.strip()
                end_seq = section.find('to_str').text.strip()
                ruling_span = section.find('ruling_span').text.strip()

                sequences = f"{start_seq} - {end_seq}"
                total_span_length = float(ruling_span) if ruling_span else 0.0
                self.output_data.append((section_num, sequences, total_span_length, circuit_type))

            self.output_data.sort(key=lambda x: int(x[0]))

            # Save the extracted span guy data to a file
            with open("extractStringingChartNeutralSpan_section_seq_totalSpanLength_circuitType.txt", "w") as file:
                max_lengths = {
                    'section_num': max(len(str(row[0])) for row in self.output_data),
                    'sequences': max(len(row[1]) for row in self.output_data),
                    'total_span_length': max(len(f"{row[2]:.2f}") for row in self.output_data),
                    'circuit_type': max(len(row[3]) for row in self.output_data)
                }
                headers = [
                    ("Section #", max_lengths['section_num']),
                    ("Sequence #s", max_lengths['sequences']),
                    ("Total Span Length", max_lengths['total_span_length']),
                    ("Circuit Type", max_lengths['circuit_type'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for row in self.output_data:
                    formatted_row = [
                        f"{row[0]:<{max_lengths['section_num']}}",
                        f"{row[1]:<{max_lengths['sequences']}}",
                        f"{row[2]:<{max_lengths['total_span_length']}.2f}",
                        f"{row[3]:<{max_lengths['circuit_type']}}"
                    ]
                    file.write(" | ".join(formatted_row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse Span Guy XML file: {e}")

    def parse_primary_conductor_data(self):
        # Parse data related to the primary conductor (main electrical wire) from the pasted data
        pasted_data = self.paste_text.get("1.0", tk.END).strip()
        if not pasted_data:
            messagebox.showerror("Error", "Please paste data into the text box.")
            return

        try:
            # Regular expression pattern to extract relevant sections from the pasted data
            sections = re.findall(r"Stringing Chart Report\n\nCircuit '(.*?)' Section #(.*?) from structure #(.*?) to structure #(.*?),.*?Span\n(.*?)\n\n", pasted_data, re.DOTALL)
            output_data = []

            for section in sections:
                circuit_type, section_num, start_seq, end_seq, spans_data = section
                spans = re.findall(r"\n\s+(\d+\.\d+)\s+", spans_data)
                total_span_length = sum(map(float, spans))
                sequences = re.findall(r"\d{4}", spans_data)  # Extract all sequences
                sequences_str = ", ".join(sequences)
                structure_to_structure = f"{start_seq} -> {end_seq}"
                circuit_value = int(re.search(r'(\d+)PH', circuit_type).group(1))
                result = total_span_length * circuit_value
                output_data.append((section_num, structure_to_structure, circuit_type, circuit_value, total_span_length, result, sequences_str))

            # Save the extracted primary conductor data to a file
            with open("extractStringingChartPrimary_section_struct_circuitType_spanLength_total.txt", "w") as file:
                max_lengths = {
                    'section_num': max(len(str(row[0])) for row in output_data),
                    'structure_to_structure': max(len(row[1]) for row in output_data),
                    'circuit_type': max(len(row[2]) for row in output_data),
                    'circuit_value': max(len(str(row[3])) for row in output_data),
                    'total_span_length': max(len(f"{row[4]:.2f}") for row in output_data),
                    'result': max(len(f"{row[5]:.2f}") for row in output_data),
                    'sequences': max(len(row[6]) for row in output_data),
                }
                headers = [
                    ("Section #", max_lengths['section_num']),
                    ("Structure -> Structure", max_lengths['structure_to_structure']),
                    ("Circuit Type", max_lengths['circuit_type']),
                    ("Circuit Value", max_lengths['circuit_value']),
                    ("Span Length", max_lengths['total_span_length']),
                    ("Result", max_lengths['result']),
                    ("Sequences", max_lengths['sequences'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for row in output_data:
                    formatted_row = [
                        f"{row[0]:<{max_lengths['section_num']}}",
                        f"{row[1]:<{max_lengths['structure_to_structure']}}",
                        f"{row[2]:<{max_lengths['circuit_type']}}",
                        f"{row[3]:<{max_lengths['circuit_value']}}",
                        f"{row[4]:<{max_lengths['total_span_length']}.2f}",
                        f"{row[5]:<{max_lengths['result']}.2f}",
                        f"{row[6]:<{max_lengths['sequences']}}",
                    ]
                    file.write(" | ".join(formatted_row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse primary conductor data: {e}")

    def parse_step6_structure_usage(self):
        # Parse XML data related to structure usage (e.g., guy wire tension) and save to file
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()
            output_data = []
            for report in root.findall('.//summary_of_maximum_element_usages_for_structure_range'):
                seq_no = report.find('str_no').text
                element_label = report.find('element_label').text
                element_type = report.find('element_type').text
                max_usage = report.find('maximum_usage').text
                if element_type == "Guy" or element_type == "Cable":
                    output_data.append((seq_no, element_label, element_type, max_usage))

            # Save the extracted structure usage data to a file
            with open("extractGuyUsage_seq_elementType_usage.txt", "w") as file:
                max_lengths = {
                    'sequence': max(len(row[0]) for row in output_data),
                    'element_label': max(len(row[1]) for row in output_data),
                    'element_type': max(len(row[2]) for row in output_data),
                    'max_usage': max(len(row[3]) for row in output_data)
                }
                headers = [
                    ("Sequence #", max_lengths['sequence']),
                    ("Element Label", max_lengths['element_label']),
                    ("Element Type", max_lengths['element_type']),
                    ("Maximum Usage", max_lengths['max_usage'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for row in output_data:
                    formatted_row = [
                        f"{row[0]:<{max_lengths['sequence']}}",
                        f"{row[1]:<{max_lengths['element_label']}}",
                        f"{row[2]:<{max_lengths['element_type']}}",
                        f"{row[3]:<{max_lengths['max_usage']}}"
                    ]
                    file.write(" | ".join(formatted_row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse XML file: {e}")

    def parse_step7_joint_support(self):
        # Parse XML data related to joint support (e.g., forces on structures) and save to file
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()
            for report in root.findall('.//summary_of_joint_support_reactions_for_all_load_cases_for_structure_range'):
                seq_no = report.find('str_no').text
                shear_force = float(report.find('shear_force').text)
                bending_moment = float(report.find('bending_moment').text)
                max_force = max(shear_force, bending_moment)

                # Store the maximum force for each sequence number
                if seq_no not in self.max_force_data:
                    self.max_force_data[seq_no] = max_force
                else:
                    self.max_force_data[seq_no] = max(self.max_force_data[seq_no], max_force)

            # Save the maximum force data to a file
            with open("extractMAX_sequence_MaxForce.txt", "w") as file:
                max_lengths = {
                    'sequence': max(len(seq) for seq in self.max_force_data),
                    'max_force': max(len(f"{force:.2f}") for force in self.max_force_data.values())
                }
                headers = [
                    ("Sequence", max_lengths['sequence']),
                    ("Max Force", max_lengths['max_force'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for seq, force in sorted(self.max_force_data.items()):
                    row = [
                        f"{seq:<{max_lengths['sequence']}}",
                        f"{force:<{max_lengths['max_force']}.2f}"
                    ]
                    file.write(" | ".join(row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse Joint Support XML file: {e}")

    def generate_report(self):
        # Generate the final report by combining data from multiple steps and saving to Excel
        # Clear previous output
        for widget in self.output_frame.winfo_children():
            widget.destroy()

        # List of files to combine into the final report
        files_to_combine = [
            "extractHIS_seq_facID_existingTrans_primaryRiser_secondaryRiser.txt",
            "extractFusingCoordination_newOrExistingFusing.txt",
            "XMLextractConstrucStakingReport_framing_type_direction_length.txt",  # Updated XML output
            "XMLextractPoleType.txt",  # Updated XML output
            "XMLextractGuyUsage_seq_elementType_usage.txt",  # Updated XML output
            "extractMAX_sequence_MaxForce.txt"
        ]

        parsed_data = {}
        for file_path in files_to_combine:
            try:
                with open(file_path, 'r') as file:
                    lines = file.readlines()
                # Parse each file and store the data in a dictionary
                if "extractHIS_seq_facID_existingTrans_primaryRiser_secondaryRiser.txt" in file_path:
                    parsed_data['his_seq'] = self.parse_his_seq(lines)
                elif "extractFusingCoordination_newOrExistingFusing.txt" in file_path:
                    parsed_data['fusing'] = self.parse_fusing_coordination(lines)
                elif "extractConstrucStakingReport_framing_type_direction_length.txt" in file_path:
                    parsed_data['construction'] = self.parse_construction_staking(lines)
                elif "extractPoleType.txt" in file_path:
                    parsed_data['pole_type'] = self.parse_pole_type(lines)
                elif "extractGuyUsage_seq_elementType_usage.txt" in file_path:
                    parsed_data['guy_usage'] = self.parse_guy_usage(lines)
                elif "extractMAX_sequence_MaxForce.txt" in file_path:
                    parsed_data['max_force'] = self.parse_max_force(lines)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read file {file_path}: {e}")
                return

        combined_data = self.combine_data(parsed_data)

        # Ask user where to save the Excel file
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            self.save_to_excel(combined_data, save_path)
            self.save_stringing_report(save_path)

    def save_stringing_report(self, file_path):
        # Save the stringing chart data (neutral and primary) to the Excel file
        neutral_span_file = "XMLextractStringingChartNeutralSpan_section_seq_totalSpanLength_circuitType.txt"
        primary_span_file = "XMLextractStringingChartPrimary_section_struct_circuitType_spanLength_total.txt"  # Updated to use XML extract file
        
        neutral_data = self.parse_stringing_file(neutral_span_file, is_primary=False)
        primary_data = self.parse_stringing_file(primary_span_file, is_primary=True)
        
        workbook = load_workbook(file_path)
        primary_sheet = workbook.create_sheet(title="Primary Stringing Data")
        neutral_sheet = workbook.create_sheet(title="Neutral Span Stringing Data")
        
        # Define headers for the data (updated for XML extract format)
        primary_headers = ["Section #", "Structure -> Structure", "Circuit Type", "Circuit Value", 
                          "Span Lengths", "Total Length", "Sequences", "Heights", "Set Numbers"]
        neutral_headers = ["Section #", "Sequence #s", "Total Span Length", "Circuit Type"]
        
        primary_sheet.append(primary_headers)
        neutral_sheet.append(neutral_headers)
        
        # Append data to the respective sheets
        for row in primary_data:
            primary_sheet.append(row)
        
        for row in neutral_data:
            neutral_sheet.append(row)
        
        # Styling for the header
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        for sheet in [primary_sheet, neutral_sheet]:
            for cell in sheet["1:1"]:
                cell.fill = header_fill
                cell.font = header_font
        
        # Adjust column widths for better readability
        for sheet in [primary_sheet, neutral_sheet]:
            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        workbook.save(file_path)
        messagebox.showinfo("Success", f"Stringing report has been added to {file_path}")

    def sequence_to_sequence_sheet(self, workbook):
        """
        Create a Sequence to Sequence sheet in the final Excel workbook
        using data from the primary stringing chart XML extract.
        """
        # Create a new sheet
        seq_to_seq_sheet = workbook.create_sheet(title="Sequence to Sequence")

        # Define headers
        headers = [
            "From Seq", 
            "To Seq", 
            "Conductor Label", 
            "Conductor Attachment Height"
        ]
        seq_to_seq_sheet.append(headers)

        # Styling for the header
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        for cell in seq_to_seq_sheet["1:1"]:
            cell.fill = header_fill
            cell.font = header_font

        # Placeholder data - you'll replace this with actual parsing later
        try:
            with open("XMLextractStringingChartPrimary_section_struct_circuitType_spanLength_total.txt", "r") as file:
                # Skip the header lines
                next(file)
                next(file)

                for line in file:
                    parts = [part.strip() for part in line.split('|')]
                    if len(parts) >= 3:
                        # Extract sequences from the "Structure -> Structure" column
                        sequences = parts[1].split('->')
                        
                        # Temporary placeholder for conductor label and height
                        conductor_label = "5R234-1/0 ACSR"
                        conductor_height = parts[7]  # Using the height column

                        for i in range(len(sequences) - 1):
                            seq_to_seq_sheet.append([
                                sequences[i].strip(), 
                                sequences[i+1].strip(), 
                                conductor_label, 
                                conductor_height.split(',')[i].strip()
                            ])

        except FileNotFoundError:
            messagebox.showwarning("File Not Found", "Primary stringing chart file not found.")
        except Exception as e:
            messagebox.showerror("Error", f"Error processing sequence to sequence sheet: {str(e)}")

        # Adjust column widths
        for column in seq_to_seq_sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            seq_to_seq_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    def parse_stringing_file(self, file_path, is_primary):
        # Parse stringing chart data from a file and return it as a list of rows
        data = []
        try:
            with open(file_path, 'r') as file:
                lines = file.readlines()[2:]  # Skip header and separator line
                for line in lines:
                    parts = [part.strip() for part in line.split('|')]
                    if is_primary and "XMLextract" in file_path:
                        # Handle the new XML-extracted primary data format
                        row_data = parts
                        data.append(row_data)
                    else:
                        # Handle neutral data or original primary data format
                        data.append(parts)
        except FileNotFoundError:
            messagebox.showwarning("File Not Found", f"Could not find {file_path}. That section will be skipped.")
        except Exception as e:
            messagebox.showerror("Error", f"Error parsing file {file_path}: {str(e)}")
            
        return data

    def save_to_excel(self, data, file_path):
        # Save the combined data from all steps into an Excel file
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data Report"

        # Add the framing sheet
        framing_sheet = workbook.create_sheet(title="Framing Report")
        self.create_framing_sheet(framing_sheet, data)
        self.sequence_to_sequence_sheet(workbook)
        
        # Define headers for the report
        headers = ['sequence', 'facility_id', 'existing_transformers', 'primary_riser', 'secondary_riser',
                   'existing_or_new_tap', 'type', 'latitude', 'longitude', 'framing', 'anchor_direction',
                   'lead_length', 'pole_type', 'element_label', 'element_type', 'max_usage', 'max_force', 'soil_class', 'description']
        sheet.append(headers)

        # Styling for the header
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        for cell in sheet["1:1"]:
            cell.fill = header_fill
            cell.font = header_font

        # Alternate row colors for better readability
        light_green = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        light_blue = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        current_fill = light_green
        previous_seq = None

        # Append data to the sheet, alternating row colors
        for seq, info in sorted(data.items(), key=lambda x: int(re.findall(r'\d+', x[0])[0])):
            max_length = max(len(info['existing_or_new_tap']), len(info['construction']), len(info['guy_usage']), 1)
            for i in range(max_length):
                row = []
                if seq != previous_seq:
                    row = [seq, info['facility_id'], info['existing_transformers'], info['primary_riser'],
                           info['secondary_riser']]
                    current_fill = light_blue if current_fill == light_green else light_green
                    previous_seq = seq
                else:
                    row = ['', '', '', '', '']

                row.append(info['existing_or_new_tap'][i] if i < len(info['existing_or_new_tap']) else '')
                if i < len(info['construction']):
                    const = info['construction'][i]
                    row.extend([const['type'], const['latitude'], const['longitude'], const['framing'],
                                const['anchor_direction'], const['lead_length']])
                else:
                    row.extend([''] * 6)
                row.append(info['pole_type'] if i == 0 else '')
                if i < len(info['guy_usage']):
                    guy = info['guy_usage'][i]
                    row.extend([guy['element_label'], guy['element_type'], guy['max_usage']])
                else:
                    row.extend([''] * 3)
                row.append(info['max_force'] if i == 0 else '')
                row.append(info.get('soil_class', ''))
                row.append(info.get('description', ''))
                sheet.append(row)

                for cell in sheet[sheet.max_row]:
                    cell.fill = current_fill

        # Adjust column widths for both sheets
        for worksheet in [sheet, framing_sheet]:
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        workbook.save(file_path)
        messagebox.showinfo("Success", f"Data has been saved to {file_path}")

    def create_framing_sheet(self, sheet, data):
        """Creates a new sheet with raw framing data"""
        # Define headers
        headers = [
            'Sequence',
            'New Framing',
            'Primary Framing',
            'Secondary Framing'
        ]
        
        # Write and style headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Process data
        row_num = 2
        for seq, info in sorted(data.items(), key=lambda x: int(re.findall(r'\d+', x[0])[0])):
            if 'construction' not in info or not info['construction']:
                continue
                
            # Get framing from construction data
            framing = info['construction'][0].get('framing', '')
            if not framing:
                continue
                
            transmission_framing = ''
            primary_framing = ''
            secondary_framing = ''
            
            # Split on '+' for primary/secondary
            parts = framing.strip().split('+')
            
            # Primary is everything before first '+'
            primary_framing = parts[0].strip()
            
            # Secondary is everything after first '+'
            if len(parts) > 1:
                secondary_framing = ' + '.join(part.strip() for part in parts[1:])
            
            # Write row data
            row_data = [
                seq,
                framing,
                primary_framing,
                secondary_framing
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col)
                cell.value = value
            
            row_num += 1


            def save_to_excel(self, data, file_path):
                # Save the combined data from all steps into an Excel file
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Data Report"

                # Add the framing sheet
                framing_sheet = workbook.create_sheet(title="Framing Report")
                self.create_framing_sheet(framing_sheet, data)

                # Add sequence to sequence sheet 
                self.sequence_to_sequence_sheet(workbook)

                # Define headers for the report
                headers = ['sequence', 'facility_id', 'existing_transformers', 'primary_riser', 'secondary_riser',
                           'existing_or_new_tap', 'type', 'latitude', 'longitude', 'framing', 'anchor_direction',
                           'lead_length', 'pole_type', 'element_label', 'element_type', 'max_usage', 'max_force', 'soil_class', 'description']
                sheet.append(headers)

                # Styling for the header
                header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                header_font = Font(bold=True)
                for cell in sheet["1:1"]:
                    cell.fill = header_fill
                    cell.font = header_font

                # Alternate row colors for better readability
                light_green = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                light_blue = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                current_fill = light_green
                previous_seq = None

                # Append data to the sheet, alternating row colors
                for seq, info in sorted(data.items(), key=lambda x: int(re.findall(r'\d+', x[0])[0])):
                    max_length = max(len(info['existing_or_new_tap']), len(info['construction']), len(info['guy_usage']), 1)
                    for i in range(max_length):
                        row = []
                        if seq != previous_seq:
                            row = [seq, info['facility_id'], info['existing_transformers'], info['primary_riser'],
                                   info['secondary_riser']]
                            current_fill = light_blue if current_fill == light_green else light_green
                            previous_seq = seq
                        else:
                            row = ['', '', '', '', '']

                        row.append(info['existing_or_new_tap'][i] if i < len(info['existing_or_new_tap']) else '')
                        if i < len(info['construction']):
                            const = info['construction'][i]
                            row.extend([const['type'], const['latitude'], const['longitude'], const['framing'],
                                        const['anchor_direction'], const['lead_length']])
                        else:
                            row.extend([''] * 6)
                        row.append(info['pole_type'] if i == 0 else '')
                        if i < len(info['guy_usage']):
                            guy = info['guy_usage'][i]
                            row.extend([guy['element_label'], guy['element_type'], guy['max_usage']])
                        else:
                            row.extend([''] * 3)
                        row.append(info['max_force'] if i == 0 else '')
                        row.append(info.get('soil_class', ''))
                        row.append(info.get('description', ''))
                        sheet.append(row)

                        for cell in sheet[sheet.max_row]:
                            cell.fill = current_fill

                # Adjust column widths for both sheets
                for worksheet in [sheet, framing_sheet]:
                    for column in worksheet.columns:
                        max_length = 0
                        column = list(column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                workbook.save(file_path)
                messagebox.showinfo("Success", f"Data has been saved to {file_path}")

    def parse_his_seq(self, lines):
        # Parse Hendrix Input Sheet (HIS) data from a file and return it as a dictionary
        data = {}
        pattern = r"(\d{4})\s+([\d\.]+|None)\s+(\d+|None)\s+(Replace|None)?\s+(Replace|None)?"
        for line in lines[1:]:
            match = re.match(pattern, line.strip())
            if match:
                seq, fac_id, existing_trans, primary_riser, secondary_riser = match.groups()
                data[seq] = {
                    'facility_id': fac_id,
                    'existing_transformers': existing_trans,
                    'primary_riser': primary_riser,
                    'secondary_riser': secondary_riser
                }
        return data

    def parse_fusing_coordination(self, lines):
        # Parse fusing coordination data from a file and return it as a dictionary
        data = {}
        pattern = r"(\d{4})\s+(.+)"
        for line in lines[1:]:
            match = re.match(pattern, line.strip())
            if match:
                seq, existing = match.groups()
                if seq not in data:
                    data[seq] = []
                data[seq].append(existing)
        return data

    def parse_construction_staking(self, lines):
        # Parse construction staking data from a file and return it as a dictionary
        data = {}
        for line in lines[2:]:  # Skip header and separator line
            parts = [part.strip() for part in line.split('|')]
            if len(parts) != 7:
                continue  # Skip lines that don't have exactly 7 parts
            seq, type_, lat, lon, framing, anchor_dir, lead_length = parts
            if seq not in data:
                data[seq] = []
            data[seq].append({
                'type': type_,
                'latitude': lat,
                'longitude': lon,
                'framing': framing,
                'anchor_direction': anchor_dir,
                'lead_length': lead_length
            })
        return data

    def parse_pole_type(self, lines):
        # Parse pole type data from a file and return it as a dictionary
        data = {}
        pattern = r"(\d{4})\s+([\w\-\.]+)"
        for line in lines[1:]:
            match = re.match(pattern, line.strip())
            if match:
                seq, pole_type = match.groups()
                data[seq] = pole_type
        return data

    def parse_guy_usage(self, lines):
        # Parse guy usage data from a file and return it as a dictionary
        data = {}
        for line in lines[2:]:  # Skip header and separator line
            parts = [part.strip() for part in line.split('|')]
            if len(parts) != 4:
                continue  # Skip lines that don't have exactly 4 parts
            seq, element_label, element_type, max_usage = parts
            if seq not in data:
                data[seq] = []
            data[seq].append({
                'element_label': element_label,
                'element_type': element_type,
                'max_usage': max_usage
            })
        return data

    def parse_max_force(self, lines):
        # Parse maximum force data from a file and return it as a dictionary
        data = {}
        for line in lines[2:]:  # Skip header and separator line
            parts = [part.strip() for part in line.split('|')]
            if len(parts) != 3:
                continue  # Skip lines that don't have exactly 3 parts
            seq, max_force, soil_class = parts
            data[seq] = {'max_force': max_force, 'soil_class': soil_class}
        return data

    def combine_data(self, parsed_data):
        # Combine parsed data from all steps into a single dictionary for final report
        combined = {}
        all_sequences = set(parsed_data['his_seq'].keys()) | set(parsed_data['fusing'].keys()) | \
                        set(parsed_data['construction'].keys()) | set(parsed_data['pole_type'].keys()) | \
                        set(parsed_data['guy_usage'].keys()) | set(parsed_data['max_force'].keys())

        for seq in all_sequences:
            combined[seq] = {
                'facility_id': parsed_data['his_seq'].get(seq, {}).get('facility_id', ''),
                'existing_transformers': parsed_data['his_seq'].get(seq, {}).get('existing_transformers', ''),
                'primary_riser': parsed_data['his_seq'].get(seq, {}).get('primary_riser', ''),
                'secondary_riser': parsed_data['his_seq'].get(seq, {}).get('secondary_riser', ''),
                'existing_or_new_tap': parsed_data['fusing'].get(seq, []),
                'construction': parsed_data['construction'].get(seq, []),
                'pole_type': parsed_data['pole_type'].get(seq, ''),
                'guy_usage': parsed_data['guy_usage'].get(seq, []),
                'max_force': parsed_data['max_force'].get(seq, {}).get('max_force', ''),
                'soil_class': parsed_data['max_force'].get(seq, {}).get('soil_class', ''),
                'description': parsed_data['max_force'].get(seq, {}).get('description', '')
            }

        return combined

if __name__ == "__main__":
    # Initialize the application and start the main loop
    root = tk.Tk()
    app = DataExtractionApp(root)
    root.mainloop()