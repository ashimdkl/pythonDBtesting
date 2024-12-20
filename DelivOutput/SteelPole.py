import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog, messagebox
import os

class SteelPoleGenerator:
    def __init__(self, workbook):
        """Initialize the generator with workbook"""
        self.workbook = workbook
        self.template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'newSteelPoleTemplate.xlsx')
        self.setup_defaults()

    def setup_defaults(self):
        """Set up default values for the steel pole information"""
        self.defaults = {
            'client_name': 'PacifiCorp',
            'project_name': '',
            'project_number': ''
        }

    def extract_pole_info(self, pole_type):
        """
        Extract pole height and class from DIST-S format pole type string
        Example: DIST-S-02.9-C1-50-0 -> 50' Class 1
        """
        if not pole_type or not isinstance(pole_type, str):
            return None

        # Pattern to match DIST-S format
        # Matches patterns like DIST-S-02.9-C1-50-0
        pattern = r'DIST-S-[\d.]+[-]C(\d+)-(\d+)-\d+'
        match = re.search(pattern, pole_type)
        
        if match:
            pole_class = match.group(1)  # Gets the class number (1)
            pole_height = match.group(2)  # Gets the height (50)
            return f"{pole_height}' Class {pole_class}"
        
        return None

    def process_pole_data(self, data_sheet):
        """Process and aggregate pole data from the input sheet"""
        pole_counts = {}  # Will store counts for each unique pole type
        
        # Iterate through rows starting from row 2 (skip header)
        for row in data_sheet.iter_rows(min_row=2, values_only=True):
            pole_type = row[13]  # Column N (14th column)
            
            if not pole_type:  # Skip empty cells
                continue

            pole_info = self.extract_pole_info(str(pole_type))
            if pole_info:
                if pole_info not in pole_counts:
                    pole_counts[pole_info] = 1
                else:
                    pole_counts[pole_info] += 1

        # Sort the pole counts by height and class
        sorted_pole_counts = {}
        for k in sorted(pole_counts.keys(), 
                       key=lambda x: (int(x.split("'")[0]), int(x.split("Class ")[1]))):
            sorted_pole_counts[k] = pole_counts[k]

        return sorted_pole_counts

    def generate_sheet(self):
        """Generate the steel pole information sheet"""
        try:
            # Load the template
            template_wb = load_workbook(self.template_path)
            sheet = template_wb.active

            # Process the pole data
            data_sheet = self.workbook.active
            pole_data = self.process_pole_data(data_sheet)

            # Start filling data at row 8
            current_row = 8
            total_count = 0

            # Fill in pole data
            for pole_type, count in pole_data.items():
                sheet.cell(row=current_row, column=1).value = pole_type
                sheet.cell(row=current_row, column=2).value = count
                total_count += count
                current_row += 1

            # Add Grand Total in column D
            sheet.cell(row=current_row, column=4).value = "Grand Total"
            sheet.cell(row=current_row, column=5).value = total_count
            sheet.cell(row=current_row, column=4).font = Font(bold=True)
            sheet.cell(row=current_row, column=5).font = Font(bold=True)

            # Save the workbook
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="Steel_Pole_Information.xlsx"
            )
            if file_path:
                template_wb.save(file_path)
                messagebox.showinfo("Success", f"File saved successfully as {file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            raise