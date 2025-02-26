import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog, messagebox

class NewFramingGenerator:
    def __init__(self, workbook):
        self.workbook = workbook
        self.setup_standards()

    def setup_standards(self):
        """Initialize the standards dictionary with all framing types"""
        self.standards = {
            'EH101': 'EH101 - 1PH TAN',
            'EH106': 'EH106 - 1PH ANGLE',
            'EH111': 'EH111 - 1PH DDE',
            'EH121': 'EH121 - 1PH CORNER',
            'EH131': 'EH131 - 1PH DE',
            'EH221': 'EH221 - 1PH TAP',
            'EH226': 'EH226 - 1PH TAP FROM 2PH LINE',
            'EH231': 'EH231 - 1PH TAP FOR FUSING',
            'EH341': 'EH341 - 1PH HORIZIONTAL DE',
            'EH421': 'EH421 - 1PH TAP WITH CROSSARM',
            'EJ300': 'EJ300 - 3PH TAN/ANGLE/DDE/DE/CORNER',
            'EJ909': 'EJ909 - 3PH TAN UNDERBUILD',
            'TF200': 'TF200 - TRANSMISSION TANGENT'
        }

    def get_standard_name(self, code_str):
        """Get full standard name from code, preserving any prefixes like (2)"""
        if not code_str:
            return ''
            
        # Extract the base standard code
        match = re.search(r'(EH|EJ|TF)\d{3}', code_str)
        if not match:
            return code_str
            
        code = match.group()
        prefix = code_str[:match.start()].strip()
        standard_name = self.standards.get(code, code)
        
        # Combine prefix with standard name if prefix exists
        return f"{prefix} {standard_name}".strip()

    def parse_framing(self, framing_str):
        """
        Parse framing string into primary and secondary components
        
        Args:
            framing_str (str): Complete framing string
            
        Returns:
            tuple: (primary_framing, secondary_framing, transmission_framing)
        """
        if not framing_str or not isinstance(framing_str, str):
            return '', '', ''

        # Check for TF200 first
        transmission_framing = ''
        if 'TF200' in framing_str:
            transmission_framing = 'TF200 - TRANSMISSION TANGENT'
            framing_str = framing_str.replace('TF200', 'EJ909')
            
        # Split on '+' to separate primary and secondary
        parts = framing_str.strip().split('+')
        
        # Primary is everything before first '+'
        primary = parts[0].strip()
        
        # Secondary is everything after first '+'
        secondary = ''
        if len(parts) > 1:
            secondary_parts = []
            for part in parts[1:]:
                # Find and convert each standard code in the secondary part
                codes = re.finditer(r'(?:\(\d+\)\s*)?(EH|EJ|TF)\d{3}', part)
                for code_match in codes:
                    code_str = code_match.group()
                    standard_name = self.get_standard_name(code_str)
                    secondary_parts.append(standard_name)
            secondary = ' + '.join(secondary_parts)
        
        # Convert primary framing to standard name
        primary = self.get_standard_name(primary)
        
        return primary, secondary, transmission_framing

    def setup_header_cell(self, cell, value):
        """Apply formatting to header cells"""
        cell.value = value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def setup_data_cell(self, cell, value):
        """Apply formatting to data cells"""
        cell.value = value
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def adjust_column_widths(self, worksheet):
        """Auto-adjust column widths based on content"""
        for col in worksheet.columns:
            max_length = 0
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    def generate_sheet(self):
        """Generate the new framing sheet"""
        try:
            # Get the first worksheet from the workbook
            data_report_sheet = self.workbook.worksheets[0]
            
            # Create new workbook and get active sheet
            output_workbook = Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "New Framing Sheet"

            # Define headers
            headers = [
                'Sequence',
                'Facility ID',
                'New Framing',
                'Transmission Framing',
                'Primary Framing Standard',
                'Secondary Framing Standard'
            ]
            
            # Setup headers
            for col_num, header in enumerate(headers, 1):
                cell = output_sheet.cell(row=1, column=col_num)
                self.setup_header_cell(cell, header)

            # Process data rows
            row_num = 2
            for row in data_report_sheet.iter_rows(min_row=2, values_only=True):
                sequence = row[0]
                if not sequence:  # Skip empty rows
                    continue

                # Extract data from row
                facility_id = row[1]
                raw_framing = row[9] if len(row) > 9 else ''  # Assuming framing is in column J

                # Parse framing into components
                primary_standard, secondary_standard, transmission_framing = self.parse_framing(raw_framing)

                # Prepare row data
                row_data = [
                    sequence,
                    facility_id,
                    raw_framing,  # Keep original framing data
                    transmission_framing,
                    primary_standard,
                    secondary_standard
                ]
                
                # Write row data
                for col_num, value in enumerate(row_data, 1):
                    cell = output_sheet.cell(row=row_num, column=col_num)
                    self.setup_data_cell(cell, value)
                
                row_num += 1

            # Adjust column widths
            self.adjust_column_widths(output_sheet)

            # Save the workbook
            self.save_workbook(output_workbook, "New_Framing_Sheet.xlsx")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while generating the framing sheet: {str(e)}")
            raise

    def save_workbook(self, workbook, default_filename):
        """Save the workbook to a user-specified location"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename
        )
        if file_path:
            try:
                workbook.save(file_path)
                messagebox.showinfo("Success", f"File saved successfully as {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save file: {str(e)}")
                raise