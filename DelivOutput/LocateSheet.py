from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tkinter import filedialog, messagebox

class LocateSheetGenerator:
    def __init__(self, workbook, wo_number, county, city_place):
        """Initialize the generator with workbook and location information"""
        self.workbook = workbook
        self.wo_number = wo_number
        self.county = county
        self.city_place = city_place
        
        # Default values for locate sheet
        self.defaults = {
            'cc_number': 11151,
            'equipment_used': 'Auger',
            'work_being_done_for': '(Utility) Pacific Power',
            'type_of_work': 'Pole Replacement',
            'directional_drilling': 'No',
            'using_equipment': 'Y',
            'within_overhead_line': 'Y',
            'location_of_work': "Locate a 30' radius around pole",
            'comments': 'Pole is marked in white paint',
            'township': '35S',
            'range': '06W',
            'section': '26',
            'quarter_section': 'SE'
        }

        # Thin border for all cells
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Light green fill for highlighted rows
        self.green_fill = PatternFill(
            start_color="CCFFCC",
            end_color="CCFFCC",
            fill_type="solid"
        )

    def setup_header_cell(self, cell, value):
        """Apply formatting to header cells (used for table headers)."""
        cell.value = value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.thin_border
        cell.fill = self.green_fill  # Table headers get green fill

    def setup_data_cell(self, cell, value):
        """Apply formatting to data cells (used for table rows)."""
        cell.value = value
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.thin_border

    def setup_info_cell(self, cell, value, fill=None, bold=False):
        """Helper to style the top info section cells."""
        cell.value = value
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.thin_border
        if fill:
            cell.fill = fill
        if bold:
            cell.font = Font(bold=True)

    def adjust_column_widths(self, worksheet):
        """Auto-adjust column widths based on content."""
        for col in worksheet.columns:
            max_length = 0
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    def get_row_data(self, sequence, facility_id, latitude_longitude=""):
        """Generate a row of data for the locate sheet table."""
        return [
            self.defaults['cc_number'],            # CC#
            self.wo_number,                        # WO#
            '',                                    # Pulling Section
            facility_id,                           # FP#
            sequence,                              # Seq#
            self.defaults['using_equipment'],      # Equipment extends 15' above ground?
            self.defaults['equipment_used'],       # Equipment Used
            self.defaults['within_overhead_line'], # Working within 10' of power line?
            self.defaults['directional_drilling'], # Directional Drilling
            self.defaults['type_of_work'],         # Type of work
            self.defaults['work_being_done_for'],  # Work being done for
            self.county,                           # County
            self.city_place,                       # City/Place
            latitude_longitude,                    # Latitude, Longitude
            self.defaults['location_of_work'],     # Location of work
            self.defaults['comments'],             # Comments
            self.defaults['township'],             # Township
            self.defaults['range'],                # Range
            self.defaults['section'],              # Section
            self.defaults['quarter_section']       # Quarter Section
        ]

    def generate_sheet(self):
        """Generate the locate sheet with contact info and a table of data."""
        try:
            # Get the first worksheet from the user's workbook
            data_report_sheet = self.workbook.worksheets[0]
            
            # Create a new workbook for the output
            output_workbook = Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "Locate Sheet"

            # ------------------------------------------------------------------
            # 1) TOP INFO SECTION (Rows 1–9)
            # ------------------------------------------------------------------

            # Row 1: "Construction Contact Information" (merged A1:F1)
            output_sheet.merge_cells('A1:F1')
            cell = output_sheet['A1']
            self.setup_info_cell(cell, "Construction Contact Information", fill=self.green_fill, bold=True)

            # Row 2: "Company:"
            self.setup_info_cell(output_sheet['A2'], "Company:")
            # Row 3: "Name:"
            self.setup_info_cell(output_sheet['A3'], "Name:")
            # Row 4: "Address:"
            self.setup_info_cell(output_sheet['A4'], "Address:")
            # Row 5: "Phone:"
            self.setup_info_cell(output_sheet['A5'], "Phone:")
            # Row 6: "Email:"
            self.setup_info_cell(output_sheet['A6'], "Email:")

            # Row 7: "Locate Contact Information" (merged A7:F7)
            output_sheet.merge_cells('A7:F7')
            cell = output_sheet['A7']
            self.setup_info_cell(cell, "Locate Contact Information", fill=self.green_fill, bold=True)

            # Row 8: "California: 1-800-642-2444 (call)"
            self.setup_info_cell(output_sheet['A8'], "California: 1-800-642-2444 (call)")
            # Row 9: "Oregon: 503-293-0826 (fax)"
            self.setup_info_cell(output_sheet['A9'], "Oregon: 503-293-0826 (fax)")

            # Leave row 10 blank
            # ------------------------------------------------------------------
            # 2) TABLE HEADERS (Row 11)
            # ------------------------------------------------------------------
            headers = [
                'CC#',
                'WO#',
                'Pulling Section',
                'FP#',
                'Seq#',
                "Will be using equipment that extends 15' above ground?",
                'Equipment Used',
                "Will you be working within 10' of an overhead power line?",
                'Directional Drilling',
                'Type of work',
                'Work being done for',
                'County',
                'City/Place',
                'Latitude, Longitude',
                'Location of work',
                'Comments',
                'Township',
                'Range',
                'Section',
                'Quarter Section'
            ]

            # Write the table headers to row 11
            for col_num, header in enumerate(headers, 1):
                cell = output_sheet.cell(row=11, column=col_num)
                self.setup_header_cell(cell, header)

            # ------------------------------------------------------------------
            # 3) TABLE DATA (starting row 12)
            # ------------------------------------------------------------------
            row_num = 12
            for row in data_report_sheet.iter_rows(min_row=2, values_only=True):
                sequence = row[0]
                if not sequence:  # Skip empty rows
                    continue

                facility_id = row[1]
                
                # Extract lat & long from columns H (index 7) and I (index 8)
                lat = row[7] if len(row) > 7 and row[7] is not None else ""
                lon = row[8] if len(row) > 8 and row[8] is not None else ""
                
                # Combine into (lat, long) string
                latitude_longitude = f"({lat}, {lon})" if lat or lon else ""

                # Generate the row data for the locate sheet
                row_data = self.get_row_data(sequence, facility_id, latitude_longitude)
                
                # Write row data to the output sheet
                for col_num, value in enumerate(row_data, 1):
                    cell = output_sheet.cell(row=row_num, column=col_num)
                    self.setup_data_cell(cell, value)
                
                row_num += 1

            # ------------------------------------------------------------------
            # 4) AUTO-ADJUST COLUMN WIDTHS
            # ------------------------------------------------------------------
            self.adjust_column_widths(output_sheet)

            # ------------------------------------------------------------------
            # 5) PROMPT USER TO SAVE THE WORKBOOK
            # ------------------------------------------------------------------
            self.save_workbook(output_workbook, "Locate_Sheet.xlsx")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while generating the locate sheet: {str(e)}")
            raise

    def save_workbook(self, workbook, default_filename):
        """Save the workbook to a user-specified location."""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename
        )
        if file_path:
            try:
                workbook.save(file_path)
                messagebox.showinfo("Success", f"File saved successfully as {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save file: {str(e)}")
                raise
