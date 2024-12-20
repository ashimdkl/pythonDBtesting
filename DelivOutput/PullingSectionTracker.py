import re
from openpyxl import load_workbook
from tkinter import filedialog, messagebox
import os

class PullingSectionTracker:
    def __init__(self, workbook, wo_number, county, city_place):
        """Initialize with workbook and user-provided information"""
        self.source_workbook = workbook
        self.wo_number = wo_number  # This comes from user input in MainUI
        self.county = county
        self.city_place = city_place
        self.template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'pullingSectionTrackerTemplate.xlsx')

    def get_sheet_by_name(self, sheet_name_pattern):
        """Get worksheet by partial name match"""
        for sheet in self.source_workbook.sheetnames:
            if sheet_name_pattern.lower() in sheet.lower():
                return self.source_workbook[sheet]
        return None

    def extract_sequences(self, sequence_str):
        """Extract and clean sequence list"""
        if not sequence_str:
            return []
        # Split by comma and handle potential ranges
        sequences = []
        parts = str(sequence_str).replace(' ', '').split(',')
        for part in parts:
            if '->' in part:  # Handle structure -> structure format
                continue
            if part:
                sequences.append(part.strip())
        return sequences

    def process_sections(self):
        """Process pulling section data from all relevant sheets"""
        primary_sheet = self.get_sheet_by_name("Primary Stringing")
        neutral_sheet = self.get_sheet_by_name("neutral span")
        main_sheet = self.source_workbook.active
        
        sections_data = {}
        
        # First pass: Collect basic section data and all possible sequences
        if primary_sheet:
            for row in primary_sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                section_num = row[0]
                from_struct = row[1] if len(row) > 1 else ""
                to_struct = row[2] if len(row) > 2 else ""
                circuit_type = row[3] if len(row) > 3 else ""
                circuit_value = row[4] if len(row) > 4 else ""
                span_length = row[5] if len(row) > 5 else 0
                result = row[6] if len(row) > 6 else None
                sequences = row[7] if len(row) > 7 else ""
                
                if section_num not in sections_data:
                    sections_data[section_num] = {
                        'work_order': self.wo_number,
                        'pull_section': f"PS{section_num}",
                        'from_seq': from_struct,
                        'to_seq': to_struct,
                        'length': span_length if span_length else 0,
                        'cable_size': f"{circuit_type} {circuit_value}".strip(),
                        'total_structures': 0,
                        'replacement_poles': 0,
                        'wire_removal': 0,
                        'sequences': set(),  # Using set to collect unique sequences
                        'sequence_list': sequences  # Store original sequence list
                    }
                else:
                    sections_data[section_num]['length'] += span_length if span_length else 0
                
                # Add sequences to the set
                seq_list = self.extract_sequences(sequences)
                if seq_list:
                    sections_data[section_num]['sequences'].update(seq_list)

        # Process main sheet for pole information
        if main_sheet:
            for row in main_sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                sequence = str(row[0])
                pole_type = row[12] if len(row) > 12 else ""  # pole_type column
                
                # Count poles for each section that contains this sequence
                for section in sections_data.values():
                    if sequence in section['sequences']:
                        section['total_structures'] += 1
                        if pole_type and 'R' in str(sequence):  # Replacement pole
                            section['replacement_poles'] += 1

        # Process neutral span data for wire removal
        if neutral_sheet:
            for row in neutral_sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                section_num = row[0]
                span_length = row[2] if len(row) > 2 else 0
                
                if section_num in sections_data:
                    sections_data[section_num]['wire_removal'] = span_length

        # Convert sets back to strings and sort
        for section in sections_data.values():
            # If we have an original sequence list, use it, otherwise use the collected sequences
            if section['sequence_list']:
                section['sequences'] = section['sequence_list']
            else:
                section['sequences'] = ', '.join(sorted(list(section['sequences'])))
            del section['sequence_list']  # Clean up temporary field

        return list(sections_data.values())

    def generate_sheet(self):
        """Generate the pulling section tracker sheet"""
        try:
            # Load template
            template_wb = load_workbook(self.template_path)
            sheet = template_wb.active

            # Process source data
            sections_data = self.process_sections()
            
            # Fill data starting below headers (row 5)
            current_row = 5
            for section in sections_data:
                # Work Order (from user input)
                sheet.cell(row=current_row, column=1).value = self.wo_number
                # Pull Section
                sheet.cell(row=current_row, column=2).value = section['pull_section']
                # From SEQ 
                sheet.cell(row=current_row, column=3).value = section['from_seq']
                # To SEQ
                sheet.cell(row=current_row, column=4).value = section['to_seq']
                # Length
                sheet.cell(row=current_row, column=5).value = section['length']
                # Cable Size
                sheet.cell(row=current_row, column=6).value = section['cable_size']
                # Total Structures
                sheet.cell(row=current_row, column=7).value = section['total_structures']
                # Replacement & New Poles
                sheet.cell(row=current_row, column=8).value = section['replacement_poles']
                # Small Wire Removal
                sheet.cell(row=current_row, column=9).value = section['wire_removal']
                # All Sequences
                sheet.cell(row=current_row, column=17).value = section['sequences']
                
                current_row += 1

            # Save the workbook
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Pulling_Section_Tracker_{self.wo_number}.xlsx"
            )
            if file_path:
                template_wb.save(file_path)
                messagebox.showinfo("Success", f"File saved successfully as {file_path}")
                return file_path

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            raise