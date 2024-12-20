import re
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from ttkthemes import ThemedTk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class DataExtractionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Extraction App")
        self.root.geometry('800x600')
        self.setup_standards()
        self.setup_gui()
        self.wo_number = None
        self.county = None
        self.city_place = None

    def setup_standards(self):
        # Update standards dictionary to match exact formatting
        self.standards = {
            'EH101': 'EH101 - 1PH TAN',
            'EH106': 'EH106 - 1PH ANGLE',
            'EH111': 'EH111 - 1PH DDE',
            'EH121': 'EH121 - 1PH CORNER',
            'EH131': 'EH131 - 1PH DE',
            'EH221': 'EH221 - 1PH TAP',
            'EH226': 'EH226 - 1PH TAP FROM 2PH LINE',
            'EH231': 'EH231 - 1PH TAP FOR FUSING',
            'EH341': 'EH341 - 1PH HORIZIONTAL DE',
            'EH421': 'EH421 - 1PH TAP WITH CROSSARM',
            'EJ300': 'EJ300 - 3PH TAN/ANGLE/DDE/DE/CORNER',
            'EJ909': 'EJ909 - 3PH TAN UNDERBUILD',
            'TF200': 'TF200 - TRANSMISSION TANGENT'
        }

    def setup_gui(self):
        self.upload_frame = ttk.Frame(self.root)
        self.upload_frame.pack(pady=20)

        self.upload_label = ttk.Label(self.upload_frame, text="Drag and Drop or Click to Upload Your Excel File", font=("Arial", 12))
        self.upload_label.pack()
        self.upload_button = ttk.Button(self.upload_frame, text="Upload Excel File", command=self.upload_file)
        self.upload_button.pack(pady=10)

        self.automation_frame = ttk.Frame(self.root)
        self.automation_frame.pack(pady=20)

        self.automation_label = ttk.Label(self.automation_frame, text="Select an Automation to Perform", font=("Arial", 14))
        self.automation_label.pack(pady=10)

        self.new_framing_button = ttk.Button(self.automation_frame, text="Generate New Framing Sheet", command=self.generate_new_framing_sheet)
        self.new_framing_button.pack(pady=5)

        self.locate_sheet_button = ttk.Button(self.automation_frame, text="Generate Locate Sheet", command=self.generate_locate_sheet)
        self.locate_sheet_button.pack(pady=5)

        self.steel_pole_button = ttk.Button(self.automation_frame, text="Generate Steel Pole Information", command=self.generate_steel_pole_information)
        self.steel_pole_button.pack(pady=5)

    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                self.workbook = load_workbook(file_path)
                self.file_path = file_path
                # Ask user for additional information in front of other windows
                self.root.lift()  # Bring main window to front
                self.wo_number = simpledialog.askstring("Input", "Please enter the WO#:", parent=self.root)
                self.county = simpledialog.askstring("Input", "Please enter the County:", parent=self.root)
                self.city_place = simpledialog.askstring("Input", "Please enter the City/Place:", parent=self.root)
                messagebox.showinfo("Success", "File uploaded successfully.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to upload file: {e}")

    def generate_new_framing_sheet(self):
        try:
            data_report_sheet = self.workbook.worksheets[0]
            output_workbook = Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "New Framing Sheet"

            headers = ['Sequence', 'Facility ID', 'New Framing', 'Transmission Framing', 
                      'Primary Framing Standard', 'Secondary Framing Standard']
            
            # Set up headers with left alignment
            for col_num, header in enumerate(headers, 1):
                cell = output_sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

            # Process data
            row_num = 2
            for row in data_report_sheet.iter_rows(min_row=2, values_only=True):
                sequence = row[0]
                if not sequence:
                    continue

                facility_id = row[1]
                primary_standard = row[9] if len(row) > 9 else None
                secondary_standard = row[10] if len(row) > 10 else None

                transmission_framing = ''
                if primary_standard and 'TF200' in str(primary_standard):
                    transmission_framing = 'TF200 - TRANSMISSION TANGENT'
                    primary_standard = 'EJ909 - 3PH TAN UNDERBUILD'

                primary_code = self.get_standard_code(primary_standard) if primary_standard else ''
                secondary_code = self.get_standard_code(secondary_standard) if secondary_standard else ''

                # Format standards
                primary_formatted = self.standards.get(primary_code, primary_standard)
                secondary_formatted = self.standards.get(secondary_code, secondary_standard)

                new_framing = self.get_new_framing(primary_standard, secondary_standard)

                # Write row data
                row_data = [sequence, facility_id, new_framing, transmission_framing, 
                           primary_formatted, secondary_formatted]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = output_sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                      right=Side(style='thin'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))
                row_num += 1

            # Auto-adjust column widths
            for col in output_sheet.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                output_sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            self.save_workbook_to_excel(output_workbook, "New_Framing_Sheet.xlsx")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def generate_locate_sheet(self):
        try:
            data_report_sheet = self.workbook.worksheets[0]
            output_workbook = Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "Locate Sheet"

            headers = ['CC#', 'WO#', 'Pulling Section', 'FP#', 'Seq#', 'Will be using equipment that extends 15\' above ground?',
                       'Equipment Used', 'Will you be working within 10\' of an overhead power line?', 'Directional Drilling',
                       'Type of work', 'Work being done for', 'County', 'City/Place', 'Latitude, Longitude', 'Location of work',
                       'Comments', 'Township', 'Range', 'Section', 'Quarter Section']
            
            # Set up headers with left alignment
            for col_num, header in enumerate(headers, 1):
                cell = output_sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

            # Process data
            row_num = 2
            for row in data_report_sheet.iter_rows(min_row=2, values_only=True):
                sequence = row[0]
                if not sequence:
                    continue

                facility_id = row[1]
                equipment_used = "Auger"  # Assuming fixed value for demonstration
                work_being_done_for = "(Utility) Pacific Power"
                type_of_work = "Pole Replacement"
                directional_drilling = "No"
                using_equipment = "Y"
                within_overhead_line = "Y"
                location_of_work = "Locate a 30' radius around pole"
                comments = "Pole is marked in white paint"
                latitude_longitude = row[13] if len(row) > 13 else ""

                # Write row data
                row_data = [11151, self.wo_number, '', facility_id, sequence, using_equipment, equipment_used,
                           within_overhead_line, directional_drilling, type_of_work, work_being_done_for, 
                           self.county, self.city_place, latitude_longitude, location_of_work, comments, '35S', '06W', '26', 'SE']
                
                for col_num, value in enumerate(row_data, 1):
                    cell = output_sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                      right=Side(style='thin'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))
                row_num += 1

            # Auto-adjust column widths
            for col in output_sheet.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                output_sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            self.save_workbook_to_excel(output_workbook, "Locate_Sheet.xlsx")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def generate_steel_pole_information(self):
        try:
            data_sheet = self.workbook.active
            output_workbook = Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "Steel Pole Information"

            # Header setup
            client_name = "PacifiCorp"
            project_name = "WO 07171392 5R234 Plumtree Ln Pt 2"
            project_number = "0245489_0000.002"

            headers = [
                "New Steel Poles", "TF200-HS-TW", "ZD017-TW", "ZD017-Stub", "Grand Total"
            ]

            # Write client/project details
            output_sheet.append(["Client:", client_name])
            output_sheet.append(["Project Name:", project_name])
            output_sheet.append(["Project Number:", project_number])
            output_sheet.append([])  # Blank row

            # Write table headers
            for col_num, header in enumerate(headers, 1):
                cell = output_sheet.cell(row=5, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Parse and aggregate data from column N (Pole Type)
            pole_data = {}
            for row in data_sheet.iter_rows(min_row=2, values_only=True):
                pole_type = row[13]  # Assuming column N is the 14th column
                if not pole_type:
                    continue

                match = re.search(r'(\d+)\'\s*Class\s*(\d+)', pole_type)
                if match:
                    height = match.group(1)
                    pole_class = match.group(2)
                    key = f"{height}' Class {pole_class}"

                    if key not in pole_data:
                        pole_data[key] = [0, 0, 0]  # TF200-HS-TW, ZD017-TW, ZD017-Stub

                    # Example logic to distribute counts (adjust as needed):
                    if "TF200" in pole_type:
                        pole_data[key][0] += 1
                    elif "TW" in pole_type:
                        pole_data[key][1] += 1
                    elif "Stub" in pole_type:
                        pole_data[key][2] += 1

            # Write pole data to the sheet
            row_num = 6
            for pole, counts in pole_data.items():
                grand_total = sum(counts)
                output_sheet.append([pole] + counts + [grand_total])
                row_num += 1

            # Add grand total row
            output_sheet.append(["Grand Total"] + [sum(x[i] for x in pole_data.values()) for i in range(3)] + [
                sum(sum(x) for x in pole_data.values())
            ])

            # Auto-adjust column widths
            for col in output_sheet.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                output_sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            self.save_workbook_to_excel(output_workbook, "Steel_Pole_Information.xlsx")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def save_workbook_to_excel(self, workbook, default_filename):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx")],
                                                initialfile=default_filename)
        if file_path:
            try:
                workbook.save(file_path)
                messagebox.showinfo("Success", f"File saved successfully as {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save file: {e}")

if __name__ == "__main__":
    root = ThemedTk(theme="arc")
    app = DataExtractionApp(root)
    root.mainloop()
