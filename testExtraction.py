import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from ttkthemes import ThemedTk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class DataExtractionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Extraction App")
        self.root.geometry('800x600')
        self.setup_standards()
        self.setup_gui()

    def setup_standards(self):
        self.standards = {
            'EH101 - 1PH TAN': '1PH TAN',
            'EH106 - 1PH ANGLE': '1PH ANGLE',
            'EH111 - 1PH DDE': '1PH DDE',
            'EH121 - 1PH CORNER': '1PH CORNER',
            'EH131 - 1PH DE': '1PH DE',
            'EH151 - 1PH DDE FOR FUSING': '1PH DDE FOR FUSING',
            'EH201 - 1PH DDE FOR LINE EXTENSION': '1PH DDE FOR LINE EXTENSION',
            'EH221 - 1PH TAP FROM 1PH LINE': '1PH TAP',
            'EH226 - 1PH TAP FROM 3PH LINE': '1PH TAP',
            'EH231 - 1PH TAP FROM 1 PH LINE FOR FUSING': '1PH TAP FOR FUSING',
            'EH236 - 1PH TAP FROM 3PH LINE FOR FUSING': '1PH TAP FOR FUSING',
            'EH301 - 1PH HORIZIONTAL TANGENT': '1PH HORIZIONTAL TANGENT',
            'EH321 - 1PH HORIZIONTAL DDE': '1PH HORIZIONTAL DDE',
            'EH331 - 1PH HORIZIONTAL CORNER': '1PH HORIZIONTAL CORNER',
            'EH341 - 1PH HORIZIONTAL DE': '1PH HORIZIONTAL DE',
            'EH401 - 1PH HORIZIONTAL DDE LINE EXTENSION': '1PH HORIZIONTAL DDE LINE EXTENSION',
            'EH421 - 1PH HORIZONTAL TAP FROM 1PH LINE': '1PH HORIZONTAL TAP FROM 1PH LINE',
            'EI101 - 2PH TAN': '2PH TAN',
            'EI131 - 2PH DDE': '2PH DDE',
            'EI141 - 2PH CORNER': '2PH CORNER',
            'EI151 - 2PH DE': '2PH DE',
            'EI171 - 2PH DDE FOR FUSING': '2PH DDE FOR FUSING',
            'EI201 - 2PH TAP ONE DIRECTION': '2PH TAP ONE DIRECTION',
            'EI221 - 2PH TAP TWO DIRECTIONS': '2PH TAP TWO DIRECTIONS',
            'EI231 - 2PH LINE EXTENSION': '2PH LINE EXTENSION',
            'EI401 - 2PH TAN HIGH NEUTRAL': '2PH TAN HIGH NEUTRAL',
            'EI431 - 2PH DDE HIGH NEUTRAL': '2PH DDE HIGH NEUTRAL',
            'EJ300 - 3PH TAN/ANGLE/DDE/DE/CORNER': '3PH TAN',
            'EJ376 - 3PH TAP FROM 3PH LINE': '3PH TAP FROM 3PH LINE',
            'EJ371 - 3PH DDE LINE EXTENSION': '3PH DDE LINE EXTENSION',
            'EJ601 - 3PH TAN ALLEY ARM': '3PH TAN ALLEY ARM',
            'EJ606 - 3PH TAN ALLEY ARM': '3PH TAN ALLEY ARM',
            'EJ621 - 3PH TAN ALLEY ARM HIGH NEUTRAL': '3PH TAN ALLEY ARM HIGH NEUTRAL',
            'EJ800 - 3PH [TAN, ANGLE, DDE, DE, CORNER]': '3PH TAN',
            'EJ861 - 3PH TAP FROM 3PH LINE': '3PH TAP FROM 3PH LINE',
            'EJ862 - 3PH TWO DIRECTIONAL TAP FROM 3PH LINE': '3PH TWO DIRECTIONAL TAP FROM 3PH LINE',
            'EJ871 - 3PH DDE LINE EXTENSION - RAPTOR AREA': '3PH DDE LINE EXTENSION - RAPTOR AREA',
            'EJ905 - 3PH TAN UNDERBUILD': '3PH TAN UNDERBUILD'
        }

        # Lookup by code (without description)
        self.code_lookup = {
            'EH101': '1PH TAN',
            'EH106': '1PH ANGLE',
            'EH111': '1PH DDE',
            'EH121': '1PH CORNER',
            'EH131': '1PH DE',
            'EH151': '1PH DDE FOR FUSING',
            'EH201': '1PH DDE FOR LINE EXTENSION',
            'EH221': '1PH TAP',
            'EH226': '1PH TAP',
            'EH231': '1PH TAP FOR FUSING',
            'EH236': '1PH TAP FOR FUSING',
            'EH301': '1PH HORIZIONTAL TANGENT',
            'EH321': '1PH HORIZIONTAL DDE',
            'EH331': '1PH HORIZIONTAL CORNER',
            'EH341': '1PH HORIZIONTAL DE',
            'EH401': '1PH HORIZIONTAL DDE LINE EXTENSION',
            'EH421': '1PH HORIZONTAL TAP FROM 1PH LINE',
            'EI101': '2PH TAN',
            'EI131': '2PH DDE',
            'EI141': '2PH CORNER',
            'EI151': '2PH DE',
            'EI171': '2PH DDE FOR FUSING',
            'EI201': '2PH TAP ONE DIRECTION',
            'EI221': '2PH TAP TWO DIRECTIONS',
            'EI231': '2PH LINE EXTENSION',
            'EI401': '2PH TAN HIGH NEUTRAL',
            'EI431': '2PH DDE HIGH NEUTRAL',
            'EJ300': '3PH TAN',
            'EJ376': '3PH TAP FROM 3PH LINE',
            'EJ371': '3PH DDE LINE EXTENSION',
            'EJ601': '3PH TAN ALLEY ARM',
            'EJ606': '3PH TAN ALLEY ARM',
            'EJ621': '3PH TAN ALLEY ARM HIGH NEUTRAL',
            'EJ800': '3PH TAN',
            'EJ861': '3PH TAP FROM 3PH LINE',
            'EJ862': '3PH TWO DIRECTIONAL TAP FROM 3PH LINE',
            'EJ871': '3PH DDE LINE EXTENSION - RAPTOR AREA',
            'EJ905': '3PH TAN UNDERBUILD'
        }

    def setup_gui(self):
        self.upload_frame = ttk.Frame(self.root)
        self.upload_frame.pack(pady=20)

        self.upload_label = ttk.Label(self.upload_frame, text="Drag and Drop or Click to Upload Your Excel File", font=("Arial", 12))
        self.upload_label.pack()
        self.upload_button = ttk.Button(self.upload_frame, text="Upload Excel File", command=self.upload_file)
        self.upload_button.pack(pady=10)

        self.automation_frame = ttk.Frame(self.root)
        self.automation_frame.pack(pady=20)

        self.automation_label = ttk.Label(self.automation_frame, text="Select an Automation to Perform", font=("Arial", 14))
        self.automation_label.pack(pady=10)

        self.new_framing_button = ttk.Button(self.automation_frame, text="Generate New Framing Sheet", command=self.generate_new_framing_sheet)
        self.new_framing_button.pack(pady=5)

    def upload_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                self.workbook = load_workbook(file_path)
                self.file_path = file_path
                messagebox.showinfo("Success", "File uploaded successfully.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to upload file: {e}")

    def get_standard_code(self, standard):
        if not standard:
            return ''
        return standard.split()[0]

    def get_new_framing(self, primary, secondary=None):
        if not primary:
            return ''

        if 'TF200' in str(primary):
            return 'TRANSMISSION TANGENT WITH 3PH TAN'

        primary_code = self.get_standard_code(primary)
        secondary_code = self.get_standard_code(secondary) if secondary else None

        # Handle base framing
        if primary_code in ['EJ300', 'EJ800']:
            if 'TAN/ANGLE' in primary:
                base = '3PH TAN/ANGLE'
            else:
                base = '3PH TAN'
        else:
            base = self.code_lookup.get(primary_code, primary)

        # Handle special combinations
        if secondary_code:
            if secondary_code == 'EH131':
                return '3PH TAN/ANGLE WITH 1PH DE'
            elif secondary_code in ['EH226', 'EH421']:
                return '3PH TAN WITH 1PH TAP'
            else:
                secondary_desc = self.code_lookup.get(secondary_code, secondary)
                return f"{base} WITH {secondary_desc}"

        return base

    def generate_new_framing_sheet(self):
        if not hasattr(self, 'workbook'):
            messagebox.showwarning("Warning", "Please upload the Excel file first.")
            return

        try:
            data_report_sheet = self.workbook.worksheets[0]
            
            output_workbook = Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "New Framing Sheet"

            # Set up headers
            headers = ['Sequence', 'Facility ID', 'New Framing', 'Transmission Framing', 
                      'Primary Framing Standard', 'Secondary Framing Standard']
            
            for col_num, header in enumerate(headers, 1):
                cell = output_sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

            # Process data and store unique sequences
            unique_data = {}
            for row in data_report_sheet.iter_rows(min_row=2, values_only=True):
                sequence = row[0]
                if sequence and sequence not in unique_data:
                    facility_id = row[1]
                    primary_standard = row[9] if len(row) > 9 else None
                    secondary_standard = row[10] if len(row) > 10 else None
                    transmission_framing = None
                    
                    if len(row) > 18:  # If description column exists
                        description = row[18]
                        if description and 'TRANSMISSION' in str(description).upper():
                            transmission_framing = 'TF200 - TRANSMISSION TANGENT'

                    # Get new framing based on standards
                    new_framing = self.get_new_framing(primary_standard, secondary_standard)

                    unique_data[sequence] = {
                        'Sequence': sequence,
                        'Facility ID': facility_id,
                        'New Framing': new_framing,
                        'Transmission Framing': transmission_framing if transmission_framing else '',
                        'Primary Framing Standard': primary_standard if primary_standard else '',
                        'Secondary Framing Standard': secondary_standard if secondary_standard else ''
                    }

            # Write unique data to output sheet
            for row_num, sequence in enumerate(sorted(unique_data.keys()), 2):
                data = unique_data[sequence]
                for col_num, key in enumerate(headers, 1):
                    cell = output_sheet.cell(row=row_num, column=col_num)
                    cell.value = data[key]
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                      right=Side(style='thin'),
                                      top=Side(style='thin'),
                                      bottom=Side(style='thin'))

            # Set column widths
            for col in range(1, 7):
                output_sheet.column_dimensions[get_column_letter(col)].width = 25

            self.save_workbook_to_excel(output_workbook, "New_Framing_Sheet.xlsx")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def save_workbook_to_excel(self, workbook, default_filename):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx")],
                                                initialfile=default_filename)
        if file_path:
            try:
                workbook.save(file_path)
                messagebox.showinfo("Success", f"File saved successfully as {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save file: {e}")

if __name__ == "__main__":
    root = ThemedTk(theme="arc")
    app = DataExtractionApp(root)
    root.mainloop()