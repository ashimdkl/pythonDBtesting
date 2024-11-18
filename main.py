import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from ttkthemes import ThemedTk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import math

class DataExtractionApp:
    def __init__(self, root):
        # Initialize the application window and set up basic variables and GUI components
        self.root = root
        self.root.title("Data Extraction App")
        self.root.geometry('1920x1080')

        # Variables to hold file paths, columns, and selected data
        self.file_path = None
        self.columns = []
        self.selected_columns = []
        self.step = 1  # Step in the process
        self.output_data = []
        self.max_force_data = {}  # Dictionary to hold maximum force data by sequence number
        self.soil_class_data = {}  # New dictionary to store soil class data

        # Set up the graphical user interface (GUI)
        self.setup_gui()

    def setup_gui(self):
        # Create the introductory frame with a button to start the analysis
        self.intro_frame = ttk.Frame(self.root)
        self.intro_frame.pack(expand=True, fill="both")
        self.start_btn = ttk.Button(self.intro_frame, text="Click to Begin Analysis!", command=self.start_analysis)
        self.start_btn.pack(pady=20)

        # Create the main frame where most of the application interactions happen
        self.main_frame = ttk.Frame(self.root)

        self.step_label = ttk.Label(self.main_frame, text="Step 1: Upload your Hendrix Input Sheet (HIS)", font=("Arial", 14))
        self.step_label.pack(pady=10)

        # Button to upload files (e.g., Excel, XML) depending on the current step
        self.upload_btn = ttk.Button(self.main_frame, text="Upload HIS Excel File", command=self.upload_file)
        self.upload_btn.pack(pady=10)

        # Listbox to allow the user to select columns from the uploaded Excel file
        self.column_label = ttk.Label(self.main_frame, text="Select Columns to Keep (including Sequence #)", font=("Arial", 14))
        self.column_label.pack(pady=10)

        # Frame and scrollbar for the listbox
        self.listbox_frame = ttk.Frame(self.main_frame)
        self.listbox_frame.pack(pady=10)
        self.column_listbox = tk.Listbox(self.listbox_frame, selectmode=tk.MULTIPLE, font=("Arial", 12), width=50, height=10)
        self.column_listbox.pack(side=tk.LEFT, fill=tk.BOTH)
        self.scrollbar = ttk.Scrollbar(self.listbox_frame, orient=tk.VERTICAL, command=self.column_listbox.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.column_listbox.config(yscrollcommand=self.scrollbar.set)

        # Buttons to process the data or skip steps
        self.process_btn = ttk.Button(self.main_frame, text="Parse Data and Move to Next Step", command=self.parse_and_next_step)
        self.process_btn.pack(pady=10)
        self.skip_btn = ttk.Button(self.main_frame, text="Skip This Step", command=self.next_step)
        self.skip_btn.pack(pady=10)

        # Widgets to paste and display data, initially hidden until needed
        self.paste_label = ttk.Label(self.main_frame, text="Paste Fusing Coordination Data Here", font=("Arial", 14))
        self.text_frame = ttk.Frame(self.main_frame)
        self.text_frame.pack(pady=10)
        self.paste_text = tk.Text(self.text_frame, wrap=tk.WORD, height=15, font=("Arial", 12))
        self.text_scrollbar = ttk.Scrollbar(self.text_frame, orient=tk.VERTICAL, command=self.paste_text.yview)
        self.paste_text.config(yscrollcommand=self.text_scrollbar.set)
        self.paste_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.text_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_frame.pack_forget()
        self.paste_label.pack_forget()

        # Button to generate the final report, hidden until the final step
        self.next_btn = ttk.Button(self.main_frame, text="Generate Report", command=self.generate_report)
        self.next_btn.pack(pady=10)
        self.next_btn.pack_forget()

        # Button for step 7 "Generate Results"
        self.gen_results_btn = ttk.Button(self.main_frame, text="Generate Deliverables", command=self.go_to_step8)
        self.gen_results_btn.pack(pady=10)
        self.gen_results_btn.pack_forget()

        # Frame to hold the output results
        self.output_frame = ttk.Frame(self.root)
        self.output_frame.pack(expand=True, fill="both")

    def start_analysis(self):
        # Transition from the intro frame to the main frame where analysis starts
        self.intro_frame.pack_forget()
        self.main_frame.pack(expand=True, fill="both")

    def upload_file(self):
        # File dialog to let the user upload a file based on the current step
        filetypes = [("Excel files", "*.xlsx *.xls")] if self.step == 1 else [("XML files", "*.xml"), ("Text files", "*.txt")]
        self.file_path = filedialog.askopenfilename(filetypes=filetypes)
        if self.file_path:
            # Call the appropriate function based on the step of the process
            if self.step == 1:
                self.load_columns_from_file()
            elif self.step == 3:
                self.parse_step3_xml()
            elif self.step == 4:
                self.parse_span_guy_xml()
            elif self.step == 6:
                self.parse_step6_structure_usage()
            elif self.step == 7:
                self.parse_step7_joint_support()
            elif self.step == 8:
                self.parse_guy_calc_report(self.file_path)
            messagebox.showinfo("File Uploaded", "File uploaded successfully.")

    def load_columns_from_file(self):
        # Load columns from the first row of the Excel file to allow user selection
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active
            self.columns = [cell.value for cell in sheet[1]]
            self.column_listbox.delete(0, tk.END)
            for column in self.columns:
                self.column_listbox.insert(tk.END, column)
        except Exception as e:
            # Show an error if something goes wrong while reading the file
            messagebox.showerror("Error", f"Failed to read file: {e}")

    def parse_data(self):
        # Parse the Excel data based on the columns selected by the user
        if not self.file_path:
            messagebox.showerror("Error", "Please upload a file first!")
            return

        # Get the columns selected by the user
        selected_indices = self.column_listbox.curselection()
        self.selected_columns = [self.column_listbox.get(i) for i in selected_indices]
        if not self.selected_columns:
            messagebox.showerror("Error", "Please select at least one column to keep!")
            return

        try:
            if self.step == 1:
                workbook = load_workbook(self.file_path)
                sheet = workbook.active

                # Open a file to write the selected data
                with open("extractHIS_seq_facID_existingTrans_primaryRiser_secondaryRiser.txt", "w") as file:
                    file.write("\t".join(self.selected_columns) + "\n")
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0] is not None:  # Assuming sequence number is the first column
                            row_data = [str(row[self.columns.index(col)]) for col in self.selected_columns]
                            file.write("\t".join(row_data) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse file: {e}")

    def next_step(self):
        # Move to the next step in the process
        self.step += 1
        self.file_path = None
        self.columns = []
        self.selected_columns = []
        self.column_listbox.delete(0, tk.END)

        # Adjust GUI elements based on the current step
        if self.step == 2:
            self.step_label.config(text="Step 2: Copy and Paste your Fusing Coordination Report")
            self.upload_btn.pack_forget()
            self.column_label.pack_forget()
            self.listbox_frame.pack_forget()
            self.process_btn.config(text="Parse Data and Move to Next Step", command=self.parse_and_next_step)
            self.paste_label.config(text="Paste Fusing Coordination Data Here")
            self.paste_label.pack(pady=10)
            self.text_frame.pack(pady=10)
            self.process_btn.pack(pady=10)
            self.skip_btn.pack(pady=10)
        elif self.step == 3:
            self.step_label.config(text="Step 3: Upload your Construction Staking Report")
            self.upload_btn.config(text="Upload Construction Staking Report")
            self.paste_label.pack_forget()
            self.text_frame.pack_forget()
            self.upload_btn.pack(pady=10)
            self.column_label.pack_forget()
            self.listbox_frame.pack_forget()
            self.process_btn.config(text="Parse Data and Move to Next Step", command=self.parse_and_next_step)
            self.process_btn.pack(pady=10)
            self.skip_btn.pack(pady=10)
        elif self.step == 4:
            self.step_label.config(text="Step 4: Copy and Paste your Stringing Chart - Neutral and Span Guy")
            self.upload_btn.pack_forget()
            self.paste_label.config(text="Paste Stringing Chart Data Here")
            self.paste_label.pack(pady=10)
            self.text_frame.pack(pady=10)
            self.process_btn.config(text="Parse Data and Move to Next Step", command=self.parse_and_next_step)
            self.process_btn.pack(pady=10)
            self.skip_btn.pack(pady=10)
        elif self.step == 5:
            self.step_label.config(text="Step 5: Copy and Paste your Stringing Chart - Primary Conductor")
            self.upload_btn.pack_forget()
            self.paste_label.config(text="Paste Primary Conductor Stringing Chart Data Here")
            self.paste_label.pack(pady=10)
            self.text_frame.pack(pady=10)
            self.process_btn.config(text="Parse Data and Move to Next Step", command=self.parse_and_next_step)
            self.process_btn.pack(pady=10)
            self.skip_btn.pack(pady=10)
        elif self.step == 6:
            self.step_label.config(text="Step 6: Upload your Structure Usage Report")
            self.upload_btn.config(text="Upload Structure Usage Report")
            self.paste_label.pack_forget()
            self.text_frame.pack_forget()
            self.upload_btn.pack(pady=10)
            self.process_btn.config(text="Parse Data and Move to Next Step", command=self.parse_and_next_step)
            self.process_btn.pack(pady=10)
            self.skip_btn.pack(pady=10)
        elif self.step == 7:
            self.step_label.config(text="Step 7: Copy and Paste Soil Class Data")
            self.paste_label.config(text="Paste Soil Class Data Here")
            self.paste_label.pack(pady=10)
            self.text_frame.pack(pady=10)
            self.upload_btn.config(text="Upload Joint Support XML and Parse")
            self.upload_btn.pack(pady=10)
            self.process_btn.config(text="Parse Soil Class Data", command=self.parse_soil_class_data)
            self.process_btn.pack(pady=10)
            self.next_btn.pack(pady=10)  # Show the Generate Report button
            self.gen_results_btn.pack(pady=10)  # Show the Generate Results button
            self.skip_btn.pack_forget()  # Remove the Skip button on the last step

    def go_to_step8(self):
        # Transition to Step 8 for generating deliverables
        self.main_frame.pack_forget()
        self.step = 8
        self.step8_frame = ttk.Frame(self.root)
        self.step8_frame.pack(expand=True, fill="both")
        self.step8_label = ttk.Label(self.step8_frame, text="Step 8: Generate Guy Calc Report Deliverable", font=("Arial", 14))
        self.step8_label.pack(pady=10)

        # Button to allow users to select the generated report Excel
        self.upload_btn_8 = ttk.Button(self.step8_frame, text="Please Select Your Generated Report", command=self.upload_generated_report)
        self.upload_btn_8.pack(pady=10)

    def upload_generated_report(self):
        # Let the user select the generated report Excel file
        filetypes = [("Excel files", "*.xlsx")]
        self.file_path = filedialog.askopenfilename(filetypes=filetypes)
        if self.file_path:
            try:
                # Load the workbook and check for available sheets
                workbook = load_workbook(self.file_path)
                available_sheets = workbook.sheetnames
                if "Data Report" in available_sheets:
                    self.parse_guy_calc_report(self.file_path)
                else:
                    messagebox.showerror("Error", "Worksheet 'Data Report' does not exist in the uploaded file.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to parse file: {e}")

    def parse_guy_calc_report(self, file_path):
        try:
            # Load the workbook and the relevant sheet
            workbook = load_workbook(file_path)
            sheet = workbook['Data Report']

            # Create a new workbook for the output
            output_workbook = Workbook()
            output_sheet = output_workbook.active
            output_sheet.title = "Guy Calc Report"

            # Add the instructions block at the top
            instructions = [
                "1. Calculations and guying specified at each location based on Final construction drawings.",
                "2. Existing and proposed pole installations are determined based on POWER Engineers PLS-CADD Plan Drawing.",
                "3. Guy label taken from PLS-CADD Pole Report provided by POWER Engineers.",
                "4. Existing and proposed guy installations are determined based on POWER Engineers EST Map.",
                "5. Guy Notes and Direction taken from POWER Engineers PLS-CADD EST Map.",
                "6. Anchor Type specified by using Anchor Selection Guide in PacifiCorp standard EG041.",
                "7. Usage % taken from 'Summary of Guy Usages - Maximum Usage %' in the PLS-CADD Pole Report provided by POWER Engineers.",
                "8. Tension is calculated using Usage % multiplied by 'Cable Properties - Ultimate Tension' in the PLS-CADD Pole Report.",
                "9. Total Max Anchor Tension is calculated by summing the guys attached to anchor, based on class 5 soil.",
                "10. All new installations are assuming 7/16\" UG guy wire installation",
                "11. Red text indicates existing Transmission guying that is assumed to be 7/16\" UG minimum.",
                "12. Guys on same pole highlighted with the same color attach to the same anchor."
            ]

            for i, instruction in enumerate(instructions, start=2):
                output_sheet.merge_cells(start_row=i, start_column=2, end_row=i, end_column=12)
                cell = output_sheet.cell(row=i, column=2, value=instruction)
                cell.alignment = Alignment(wrap_text=True)

            # Add headers
            headers = [
                "Pole Number", "Facility Pole ID", "Existing or Proposed Pole", "Guy Label", "Type of Guy",
                "Lead Length", "Anchor Direction", "Anchor Number", "Anchor Type", "% Usage", "Guy Tension (lb)",
                "Total Max Anchor Tension (lb)", "Material to Include in RCMS Estimate"
            ]
            header_row = len(instructions) + 3  # Start headers after instructions
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_num, header in enumerate(headers, start=2):
                cell = output_sheet.cell(row=header_row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Append the data from the 'Data Report' sheet
            for row in sheet.iter_rows(min_row=2, values_only=True):
                type_of_guy = row[6] if row[6] is not None else ""
                
                # Logic to extract proper guy label from connections
                if "to" in type_of_guy:  # Assuming column 6 contains the type like 'P1 to PG11'
                    guy_label = type_of_guy.split("to")[1].strip()
                    output_sheet.append([
                        row[0],  # Pole Number (sequence)
                        row[1],  # Facility Pole ID (facility_id)
                        "Proposed" if row[2] else "Existing",  # Existing or Proposed Pole based on transformers
                        guy_label,  # Extracted Guy Label
                        type_of_guy,  # Type of Guy (from type column)
                        row[8],  # Lead Length
                        row[7],  # Anchor Direction
                        None,    # Placeholder for Anchor Number (add logic if required)
                        row[9],  # Anchor Type
                        row[10], # % Usage
                        row[11], # Guy Tension
                        row[12], # Total Max Anchor Tension
                        None     # Placeholder for Material to Include in RCMS Estimate (add logic if required)
                    ])

            # Auto adjust column widths for readability
            for col in output_sheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                output_sheet.column_dimensions[col_letter].width = adjusted_width

            # Prompt the user to save the final report
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if save_path:
                output_workbook.save(save_path)
                messagebox.showinfo("Success", f"Guy Calc Report saved to {save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Guy Calc Report: {e}")

    def parse_and_next_step(self):
        # Parse the data and move to the next step based on the current step
        if self.step == 1:
            self.parse_data()
        elif self.step == 2:
            self.parse_pasted_data()
        elif self.step == 3:
            self.parse_step3_xml()
        elif self.step == 4:
            self.parse_and_continue_stringing_chart()
        elif self.step == 5:
            self.parse_primary_conductor_data()
        elif self.step == 6:
            self.parse_step6_structure_usage()
        elif self.step == 7:
            self.parse_step7_joint_support()
        self.next_step()

    def parse_pasted_data(self):
        # Parse data pasted by the user into the text box
        pasted_data = self.paste_text.get("1.0", tk.END).strip()
        if not pasted_data:
            messagebox.showerror("Error", "Please paste data into the text box.")
            return

        try:
            # Split the pasted data into lines and extract relevant fields using regular expressions
            lines = pasted_data.split("\n")
            header = lines[0].split("\t")
            data = []
            for line in lines[1:]:
                fields = line.split("\t")
                sequence = fields[0][:4]  # Assuming sequence is the first 4 characters
                existing = fields[2]  # Assuming existing value is the second field
                data.append([sequence, existing])

            # Save the extracted data to a file
            with open("extractFusingCoordination_newOrExistingFusing.txt", "w") as file:
                file.write("Sequence\tExisting\n")
                for row in data:
                    file.write("\t".join(row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse pasted data: {e}")

    def parse_soil_class_data(self):
        # Parse soil class data pasted by the user into the text box
        pasted_data = self.paste_text.get("1.0", tk.END).strip()
        if not pasted_data:
            messagebox.showerror("Error", "Please paste data into the text box.")
            return

        try:
            # Regular expression pattern to extract soil class information from the pasted data
            lines = pasted_data.split('\n')
            pattern = r'(\d+)(?:-(\d+))?\s+(\d+)\s+(.*)'
            
            for line in lines:
                match = re.match(pattern, line.strip())
                if match:
                    start_seq, end_seq, soil_class, description = match.groups()
                    start_seq = int(start_seq)
                    end_seq = int(end_seq) if end_seq else start_seq

                    # Store the soil class information for each sequence in the range
                    for seq in range(start_seq, end_seq + 1):
                        self.soil_class_data[seq] = {'soil_class': soil_class, 'description': description}

            messagebox.showinfo("Success", "Soil class data parsed successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse soil class data: {e}")

        # Now, update the MAX force file with soil class information
        self.update_max_force_file_with_soil_class()

    def update_max_force_file_with_soil_class(self):
        # Update the MAX force file with the soil class data parsed earlier
        try:
            with open("extractMAX_sequence_MaxForce.txt", "r") as file:
                lines = file.readlines()

            updated_lines = [lines[0], "Sequence | Max Force | Soil Class\n", "-" * 40 + "\n"]

            for line in lines[2:]:
                parts = line.strip().split('|')
                sequence = int(parts[0].strip())
                max_force = parts[1].strip()
                soil_class = self.soil_class_data.get(sequence, {}).get('soil_class', 'N/A')
                updated_lines.append(f"{sequence:4d} | {max_force:8} | {soil_class}\n")

            # Write the updated lines back to the file
            with open("extractMAX_sequence_MaxForce.txt", "w") as file:
                file.writelines(updated_lines)

            messagebox.showinfo("Success", "MAX force file updated with soil class information.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update MAX force file: {e}")

    def parse_step3_xml(self):
        # Parse the XML file for construction staking data
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()
            data = {}
            pole_types = {}

            # Extract relevant data from the XML structure
            for report in root.findall('.//construction_staking_report'):
                sequence = report.find('structure_number').text or ''
                framing = report.find('structure_name').text or ''
                latitude = report.find('latitude').text or ''
                longitude = report.find('longitude').text or ''
                x_easting = report.find('x_easting').text or ''
                y_northing = report.find('y_northing').text or ''
                stake_description = report.find('stake_description').text or ''

                # Store pole type if it is "P1"
                if "P1" in stake_description:
                    pole_type = report.find('pole_property_label').text or ''
                    pole_types[sequence] = pole_type

                # Simplify framing description by removing extraneous parts
                framing_parts = framing.split(" ", 2)
                if len(framing_parts) > 2:
                    framing = framing_parts[-1]
                    framing = " ".join(framing.split()[:-1])

                if sequence not in data:
                    data[sequence] = []

                # Store the parsed data in a dictionary
                data[sequence].append({
                    'framing': framing,
                    'latitude': latitude,
                    'longitude': longitude,
                    'x_easting': x_easting,
                    'y_northing': y_northing,
                    'stake_description': stake_description
                })

            anchor_data = []
            guy_types = ["P2", "PG", "SE", "NG", "CM", "FG"]
            for sequence, points in data.items():
                p1_point = None
                for point in points:
                    if "P1" in point['stake_description']:
                        p1_point = point
                        break

                # Calculate lead length and direction for anchor points
                if p1_point:
                    x_origin = float(p1_point['x_easting'])
                    y_origin = float(p1_point['y_northing'])
                    stake_description_set = set()
                    for point in points:
                        for guy_type in guy_types:
                            if guy_type in point['stake_description']:
                                x_next = float(point['x_easting'])
                                y_next = float(point['y_northing'])
                                lead_length = math.sqrt((x_next - x_origin) ** 2 + (y_next - y_origin) ** 2)
                                theta = math.degrees(math.atan2(y_next - y_origin, x_next - x_origin))
                                direction = self.get_cardinal_direction(theta)
                                descriptions = point['stake_description'].split(',')
                                for description in descriptions:
                                    if description.strip() not in stake_description_set:
                                        stake_description_set.add(description.strip())
                                        anchor_data.append({
                                            'sequence': sequence,
                                            'type': f"P1 to {description.strip()}",
                                            'latitude': point['latitude'],
                                            'longitude': point['longitude'],
                                            'framing': point['framing'],
                                            'anchor_direction': direction,
                                            'lead_length': lead_length
                                        })

            anchor_data.sort(key=lambda x: x['sequence'])

            # Save the extracted anchor data to a file
            with open("extractConstrucStakingReport_framing_type_direction_length.txt", "w") as file:
                max_lengths = {
                    'sequence': max(len(item['sequence']) for item in anchor_data),
                    'type': max(len(item['type']) for item in anchor_data),
                    'latitude': max(len(str(item['latitude'])) for item in anchor_data),
                    'longitude': max(len(str(item['longitude'])) for item in anchor_data),
                    'framing': max(len(item['framing']) for item in anchor_data),
                    'anchor_direction': max(len(item['anchor_direction']) for item in anchor_data),
                    'lead_length': max(len(f"{item['lead_length']:.2f}") for item in anchor_data)
                }
                headers = [
                    ("Sequence", max_lengths['sequence']),
                    ("Type", max_lengths['type']),
                    ("Latitude", max_lengths['latitude']),
                    ("Longitude", max_lengths['longitude']),
                    ("Framing", max_lengths['framing']),
                    ("Anchor Direction", max_lengths['anchor_direction']),
                    ("Lead Length", max_lengths['lead_length'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for item in anchor_data:
                    row = [
                        f"{item['sequence']:<{max_lengths['sequence']}}",
                        f"{item['type']:<{max_lengths['type']}}",
                        f"{item['latitude']:<{max_lengths['latitude']}}",
                        f"{item['longitude']:<{max_lengths['longitude']}}",
                        f"{item['framing']:<{max_lengths['framing']}}",
                        f"{item['anchor_direction']:<{max_lengths['anchor_direction']}}",
                        f"{item['lead_length']:<{max_lengths['lead_length']}.2f}"
                    ]
                    file.write(" | ".join(row) + "\n")

            # Save pole type data to a separate file
            with open("extractPoleType.txt", "w") as file:
                file.write("Sequence\tPole Type\n")
                for sequence, pole_type in sorted(pole_types.items()):
                    file.write(f"{sequence}\t{pole_type}\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse XML file: {e}")

    def get_cardinal_direction(self, angle):
        # Convert an angle into a cardinal direction (e.g., N, NE, E)
        if -2.0 < angle <= 2.0:
            return 'E'
        elif 2.0 < angle <= 88.0:
            return 'NE'
        elif 88.0 < angle <= 92.0:
            return 'N'
        elif 92.0 < angle <= 178.0:
            return 'NW'
        elif -88.0 < angle <= -2.0:
            return 'SE'
        elif -92.0 < angle <= -88.0:
            return 'S'
        elif -178.0 < angle <= -92.0:
            return 'SW'
        else:
            return 'W'

    def parse_and_continue_stringing_chart(self):
        # Parse stringing chart data pasted by the user and continue processing with the XML file
        pasted_data = self.paste_text.get("1.0", tk.END).strip()
        if not pasted_data:
            messagebox.showerror("Error", "Please paste data into the text box.")
            return

        try:
            # Use regular expressions to extract relevant sections from the pasted data
            sections = re.findall(r"Stringing Chart Report\n\nCircuit '(.*?)' Section #(.*?) from structure #(.*?) to structure #(.*?),.*?Span\n(.*?)\n\n", pasted_data, re.DOTALL)
            self.output_data = []

            for section in sections:
                circuit_type, section_num, start_seq, end_seq, spans_data = section
                if "Span Guy" in circuit_type:
                    continue  # Skip Span Guy entries from the pasted data

                spans = re.findall(r"\n\s+(\d+\.\d+)\s+", spans_data)
                total_span_length = sum(map(float, spans))
                sequences = f"{start_seq} - {end_seq}"
                self.output_data.append((section_num, sequences, total_span_length, circuit_type))

            messagebox.showinfo("Success", f"Pasted data parsed successfully. Please upload the Span Guy XML file.")
            self.upload_file()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse stringing chart data: {e}")

    def parse_span_guy_xml(self):
        # Parse the XML file for span guy data (wires supporting poles)
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()

            for section in root.findall('.//section_sagging_data'):
                circuit_type = section.find('circuit').text.strip()
                section_num = section.find('sec_no').text.strip()
                start_seq = section.find('from_str').text.strip()
                end_seq = section.find('to_str').text.strip()
                ruling_span = section.find('ruling_span').text.strip()

                sequences = f"{start_seq} - {end_seq}"
                total_span_length = float(ruling_span) if ruling_span else 0.0
                self.output_data.append((section_num, sequences, total_span_length, circuit_type))

            self.output_data.sort(key=lambda x: int(x[0]))

            # Save the extracted span guy data to a file
            with open("extractStringingChartNeutralSpan_section_seq_totalSpanLength_circuitType.txt", "w") as file:
                max_lengths = {
                    'section_num': max(len(str(row[0])) for row in self.output_data),
                    'sequences': max(len(row[1]) for row in self.output_data),
                    'total_span_length': max(len(f"{row[2]:.2f}") for row in self.output_data),
                    'circuit_type': max(len(row[3]) for row in self.output_data)
                }
                headers = [
                    ("Section #", max_lengths['section_num']),
                    ("Sequence #s", max_lengths['sequences']),
                    ("Total Span Length", max_lengths['total_span_length']),
                    ("Circuit Type", max_lengths['circuit_type'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for row in self.output_data:
                    formatted_row = [
                        f"{row[0]:<{max_lengths['section_num']}}",
                        f"{row[1]:<{max_lengths['sequences']}}",
                        f"{row[2]:<{max_lengths['total_span_length']}.2f}",
                        f"{row[3]:<{max_lengths['circuit_type']}}"
                    ]
                    file.write(" | ".join(formatted_row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse Span Guy XML file: {e}")

    def parse_primary_conductor_data(self):
        # Parse data related to the primary conductor (main electrical wire) from the pasted data
        pasted_data = self.paste_text.get("1.0", tk.END).strip()
        if not pasted_data:
            messagebox.showerror("Error", "Please paste data into the text box.")
            return

        try:
            # Regular expression pattern to extract relevant sections from the pasted data
            sections = re.findall(r"Stringing Chart Report\n\nCircuit '(.*?)' Section #(.*?) from structure #(.*?) to structure #(.*?),.*?Span\n(.*?)\n\n", pasted_data, re.DOTALL)
            output_data = []

            for section in sections:
                circuit_type, section_num, start_seq, end_seq, spans_data = section
                spans = re.findall(r"\n\s+(\d+\.\d+)\s+", spans_data)
                total_span_length = sum(map(float, spans))
                sequences = re.findall(r"\d{4}", spans_data)  # Extract all sequences
                sequences_str = ", ".join(sequences)
                structure_to_structure = f"{start_seq} -> {end_seq}"
                circuit_value = int(re.search(r'(\d+)PH', circuit_type).group(1))
                result = total_span_length * circuit_value
                output_data.append((section_num, structure_to_structure, circuit_type, circuit_value, total_span_length, result, sequences_str))

            # Save the extracted primary conductor data to a file
            with open("extractStringingChartPrimary_section_struct_circuitType_spanLength_total.txt", "w") as file:
                max_lengths = {
                    'section_num': max(len(str(row[0])) for row in output_data),
                    'structure_to_structure': max(len(row[1]) for row in output_data),
                    'circuit_type': max(len(row[2]) for row in output_data),
                    'circuit_value': max(len(str(row[3])) for row in output_data),
                    'total_span_length': max(len(f"{row[4]:.2f}") for row in output_data),
                    'result': max(len(f"{row[5]:.2f}") for row in output_data),
                    'sequences': max(len(row[6]) for row in output_data),
                }
                headers = [
                    ("Section #", max_lengths['section_num']),
                    ("Structure -> Structure", max_lengths['structure_to_structure']),
                    ("Circuit Type", max_lengths['circuit_type']),
                    ("Circuit Value", max_lengths['circuit_value']),
                    ("Span Length", max_lengths['total_span_length']),
                    ("Result", max_lengths['result']),
                    ("Sequences", max_lengths['sequences'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for row in output_data:
                    formatted_row = [
                        f"{row[0]:<{max_lengths['section_num']}}",
                        f"{row[1]:<{max_lengths['structure_to_structure']}}",
                        f"{row[2]:<{max_lengths['circuit_type']}}",
                        f"{row[3]:<{max_lengths['circuit_value']}}",
                        f"{row[4]:<{max_lengths['total_span_length']}.2f}",
                        f"{row[5]:<{max_lengths['result']}.2f}",
                        f"{row[6]:<{max_lengths['sequences']}}",
                    ]
                    file.write(" | ".join(formatted_row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse primary conductor data: {e}")

    def parse_step6_structure_usage(self):
        # Parse XML data related to structure usage (e.g., guy wire tension) and save to file
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()
            output_data = []
            for report in root.findall('.//summary_of_maximum_element_usages_for_structure_range'):
                seq_no = report.find('str_no').text
                element_label = report.find('element_label').text
                element_type = report.find('element_type').text
                max_usage = report.find('maximum_usage').text
                if element_type == "Guy" or element_type == "Cable":
                    output_data.append((seq_no, element_label, element_type, max_usage))

            # Save the extracted structure usage data to a file
            with open("extractGuyUsage_seq_elementType_usage.txt", "w") as file:
                max_lengths = {
                    'sequence': max(len(row[0]) for row in output_data),
                    'element_label': max(len(row[1]) for row in output_data),
                    'element_type': max(len(row[2]) for row in output_data),
                    'max_usage': max(len(row[3]) for row in output_data)
                }
                headers = [
                    ("Sequence #", max_lengths['sequence']),
                    ("Element Label", max_lengths['element_label']),
                    ("Element Type", max_lengths['element_type']),
                    ("Maximum Usage", max_lengths['max_usage'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for row in output_data:
                    formatted_row = [
                        f"{row[0]:<{max_lengths['sequence']}}",
                        f"{row[1]:<{max_lengths['element_label']}}",
                        f"{row[2]:<{max_lengths['element_type']}}",
                        f"{row[3]:<{max_lengths['max_usage']}}"
                    ]
                    file.write(" | ".join(formatted_row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse XML file: {e}")

    def parse_step7_joint_support(self):
        # Parse XML data related to joint support (e.g., forces on structures) and save to file
        try:
            tree = ET.parse(self.file_path)
            root = tree.getroot()
            for report in root.findall('.//summary_of_joint_support_reactions_for_all_load_cases_for_structure_range'):
                seq_no = report.find('str_no').text
                shear_force = float(report.find('shear_force').text)
                bending_moment = float(report.find('bending_moment').text)
                max_force = max(shear_force, bending_moment)

                # Store the maximum force for each sequence number
                if seq_no not in self.max_force_data:
                    self.max_force_data[seq_no] = max_force
                else:
                    self.max_force_data[seq_no] = max(self.max_force_data[seq_no], max_force)

            # Save the maximum force data to a file
            with open("extractMAX_sequence_MaxForce.txt", "w") as file:
                max_lengths = {
                    'sequence': max(len(seq) for seq in self.max_force_data),
                    'max_force': max(len(f"{force:.2f}") for force in self.max_force_data.values())
                }
                headers = [
                    ("Sequence", max_lengths['sequence']),
                    ("Max Force", max_lengths['max_force'])
                ]

                header_row = " | ".join(f"{header[0]:<{header[1]}}" for header in headers)
                file.write(header_row + "\n")
                file.write("-" * len(header_row) + "\n")

                for seq, force in sorted(self.max_force_data.items()):
                    row = [
                        f"{seq:<{max_lengths['sequence']}}",
                        f"{force:<{max_lengths['max_force']}.2f}"
                    ]
                    file.write(" | ".join(row) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse Joint Support XML file: {e}")

    def generate_report(self):
        # Generate the final report by combining data from multiple steps and saving to Excel
        # Clear previous output
        for widget in self.output_frame.winfo_children():
            widget.destroy()

        # List of files to combine into the final report
        files_to_combine = [
            "extractHIS_seq_facID_existingTrans_primaryRiser_secondaryRiser.txt",
            "extractFusingCoordination_newOrExistingFusing.txt",
            "extractConstrucStakingReport_framing_type_direction_length.txt",
            "extractPoleType.txt",
            "extractGuyUsage_seq_elementType_usage.txt",
            "extractMAX_sequence_MaxForce.txt"
        ]

        parsed_data = {}
        for file_path in files_to_combine:
            try:
                with open(file_path, 'r') as file:
                    lines = file.readlines()
                # Parse each file and store the data in a dictionary
                if "extractHIS_seq_facID_existingTrans_primaryRiser_secondaryRiser.txt" in file_path:
                    parsed_data['his_seq'] = self.parse_his_seq(lines)
                elif "extractFusingCoordination_newOrExistingFusing.txt" in file_path:
                    parsed_data['fusing'] = self.parse_fusing_coordination(lines)
                elif "extractConstrucStakingReport_framing_type_direction_length.txt" in file_path:
                    parsed_data['construction'] = self.parse_construction_staking(lines)
                elif "extractPoleType.txt" in file_path:
                    parsed_data['pole_type'] = self.parse_pole_type(lines)
                elif "extractGuyUsage_seq_elementType_usage.txt" in file_path:
                    parsed_data['guy_usage'] = self.parse_guy_usage(lines)
                elif "extractMAX_sequence_MaxForce.txt" in file_path:
                    parsed_data['max_force'] = self.parse_max_force(lines)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read file {file_path}: {e}")
                return

        combined_data = self.combine_data(parsed_data)

        # Ask user where to save the Excel file
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            self.save_to_excel(combined_data, save_path)
            self.save_stringing_report(save_path)

    def save_stringing_report(self, file_path):
        # Save the stringing chart data (neutral and primary) to the Excel file
        neutral_span_file = "extractStringingChartNeutralSpan_section_seq_totalSpanLength_circuitType.txt"
        primary_span_file = "extractStringingChartPrimary_section_struct_circuitType_spanLength_total.txt"
        
        neutral_data = self.parse_stringing_file(neutral_span_file, is_primary=False)
        primary_data = self.parse_stringing_file(primary_span_file, is_primary=True)
        
        workbook = load_workbook(file_path)
        primary_sheet = workbook.create_sheet(title="Primary Stringing Data")
        neutral_sheet = workbook.create_sheet(title="Neutral Span Stringing Data")
        
        # Define headers for the data
        primary_headers = ["Section #", "Structure -> Structure", "Circuit Type", "Circuit Value", "Span Length", "Result", "Sequences"]
        neutral_headers = ["Section #", "Sequence #s", "Total Span Length", "Circuit Type"]
        
        primary_sheet.append(primary_headers)
        neutral_sheet.append(neutral_headers)
        
        # Append data to the respective sheets
        for row in primary_data:
            primary_sheet.append(row)
        
        for row in neutral_data:
            neutral_sheet.append(row)
        
        # Styling for the header
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        for sheet in [primary_sheet, neutral_sheet]:
            for cell in sheet["1:1"]:
                cell.fill = header_fill
                cell.font = header_font
        
        # Adjust column widths for better readability
        for sheet in [primary_sheet, neutral_sheet]:
            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        workbook.save(file_path)
        messagebox.showinfo("Success", f"Stringing report has been added to {file_path}")

    def parse_stringing_file(self, file_path, is_primary):
        # Parse stringing chart data from a file and return it as a list of rows
        data = []
        with open(file_path, 'r') as file:
            lines = file.readlines()[2:]  # Skip header and separator line
            for line in lines:
                parts = [part.strip() for part in line.split('|')]
                if is_primary:
                    data.append(parts)
                else:
                    data.append(parts)
        return data

    def save_to_excel(self, data, file_path):
        # Save the combined data from all steps into an Excel file
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data Report"

        # Define headers for the report
        headers = ['sequence', 'facility_id', 'existing_transformers', 'primary_riser', 'secondary_riser',
                   'existing_or_new_tap', 'type', 'latitude', 'longitude', 'framing', 'anchor_direction',
                   'lead_length', 'pole_type', 'element_label', 'element_type', 'max_usage', 'max_force', 'soil_class', 'description']
        sheet.append(headers)

        # Styling for the header
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        for cell in sheet["1:1"]:
            cell.fill = header_fill
            cell.font = header_font

        # Alternate row colors for better readability
        light_green = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        light_blue = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        current_fill = light_green
        previous_seq = None

        # Append data to the sheet, alternating row colors
        for seq, info in sorted(data.items(), key=lambda x: int(re.findall(r'\d+', x[0])[0])):
            max_length = max(len(info['existing_or_new_tap']), len(info['construction']), len(info['guy_usage']), 1)
            for i in range(max_length):
                row = []
                if seq != previous_seq:
                    row = [seq, info['facility_id'], info['existing_transformers'], info['primary_riser'],
                           info['secondary_riser']]
                    # Alternate fill color when sequence changes
                    current_fill = light_blue if current_fill == light_green else light_green
                    previous_seq = seq
                else:
                    row = ['', '', '', '', '']  # Leave sequence and related fields blank

                row.append(info['existing_or_new_tap'][i] if i < len(info['existing_or_new_tap']) else '')
                if i < len(info['construction']):
                    const = info['construction'][i]
                    row.extend([const['type'], const['latitude'], const['longitude'], const['framing'],
                                const['anchor_direction'], const['lead_length']])
                else:
                    row.extend([''] * 6)
                row.append(info['pole_type'] if i == 0 else '')
                if i < len(info['guy_usage']):
                    guy = info['guy_usage'][i]
                    row.extend([guy['element_label'], guy['element_type'], guy['max_usage']])
                else:
                    row.extend([''] * 3)
                row.append(info['max_force'] if i == 0 else '')
                row.append(info.get('soil_class', ''))  # Add soil class
                row.append(info.get('description', ''))  # Add soil class description
                sheet.append(row)

                # Apply the current fill color to the row
                for cell in sheet[sheet.max_row]:
                    cell.fill = current_fill

        # Adjust column widths for better readability
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                        pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        workbook.save(file_path)
        messagebox.showinfo("Success", f"Data has been saved to {file_path}")

    def parse_his_seq(self, lines):
        # Parse Hendrix Input Sheet (HIS) data from a file and return it as a dictionary
        data = {}
        pattern = r"(\d{4})\s+([\d\.]+|None)\s+(\d+|None)\s+(Replace|None)?\s+(Replace|None)?"
        for line in lines[1:]:
            match = re.match(pattern, line.strip())
            if match:
                seq, fac_id, existing_trans, primary_riser, secondary_riser = match.groups()
                data[seq] = {
                    'facility_id': fac_id,
                    'existing_transformers': existing_trans,
                    'primary_riser': primary_riser,
                    'secondary_riser': secondary_riser
                }
        return data

    def parse_fusing_coordination(self, lines):
        # Parse fusing coordination data from a file and return it as a dictionary
        data = {}
        pattern = r"(\d{4})\s+(.+)"
        for line in lines[1:]:
            match = re.match(pattern, line.strip())
            if match:
                seq, existing = match.groups()
                if seq not in data:
                    data[seq] = []
                data[seq].append(existing)
        return data

    def parse_construction_staking(self, lines):
        # Parse construction staking data from a file and return it as a dictionary
        data = {}
        for line in lines[2:]:  # Skip header and separator line
            parts = [part.strip() for part in line.split('|')]
            if len(parts) != 7:
                continue  # Skip lines that don't have exactly 7 parts
            seq, type_, lat, lon, framing, anchor_dir, lead_length = parts
            if seq not in data:
                data[seq] = []
            data[seq].append({
                'type': type_,
                'latitude': lat,
                'longitude': lon,
                'framing': framing,
                'anchor_direction': anchor_dir,
                'lead_length': lead_length
            })
        return data

    def parse_pole_type(self, lines):
        # Parse pole type data from a file and return it as a dictionary
        data = {}
        pattern = r"(\d{4})\s+([\w\-\.]+)"
        for line in lines[1:]:
            match = re.match(pattern, line.strip())
            if match:
                seq, pole_type = match.groups()
                data[seq] = pole_type
        return data

    def parse_guy_usage(self, lines):
        # Parse guy usage data from a file and return it as a dictionary
        data = {}
        for line in lines[2:]:  # Skip header and separator line
            parts = [part.strip() for part in line.split('|')]
            if len(parts) != 4:
                continue  # Skip lines that don't have exactly 4 parts
            seq, element_label, element_type, max_usage = parts
            if seq not in data:
                data[seq] = []
            data[seq].append({
                'element_label': element_label,
                'element_type': element_type,
                'max_usage': max_usage
            })
        return data

    def parse_max_force(self, lines):
        # Parse maximum force data from a file and return it as a dictionary
        data = {}
        for line in lines[2:]:  # Skip header and separator line
            parts = [part.strip() for part in line.split('|')]
            if len(parts) != 3:
                continue  # Skip lines that don't have exactly 3 parts
            seq, max_force, soil_class = parts
            data[seq] = {'max_force': max_force, 'soil_class': soil_class}
        return data

    def combine_data(self, parsed_data):
        # Combine parsed data from all steps into a single dictionary for final report
        combined = {}
        all_sequences = set(parsed_data['his_seq'].keys()) | set(parsed_data['fusing'].keys()) | \
                        set(parsed_data['construction'].keys()) | set(parsed_data['pole_type'].keys()) | \
                        set(parsed_data['guy_usage'].keys()) | set(parsed_data['max_force'].keys())

        for seq in all_sequences:
            combined[seq] = {
                'facility_id': parsed_data['his_seq'].get(seq, {}).get('facility_id', ''),
                'existing_transformers': parsed_data['his_seq'].get(seq, {}).get('existing_transformers', ''),
                'primary_riser': parsed_data['his_seq'].get(seq, {}).get('primary_riser', ''),
                'secondary_riser': parsed_data['his_seq'].get(seq, {}).get('secondary_riser', ''),
                'existing_or_new_tap': parsed_data['fusing'].get(seq, []),
                'construction': parsed_data['construction'].get(seq, []),
                'pole_type': parsed_data['pole_type'].get(seq, ''),
                'guy_usage': parsed_data['guy_usage'].get(seq, []),
                'max_force': parsed_data['max_force'].get(seq, {}).get('max_force', ''),
                'soil_class': parsed_data['max_force'].get(seq, {}).get('soil_class', ''),
                'description': parsed_data['max_force'].get(seq, {}).get('description', '')
            }

        return combined

if __name__ == "__main__":
    # Initialize the application and start the main loop
    root = tk.Tk()
    app = DataExtractionApp(root)
    root.mainloop()
