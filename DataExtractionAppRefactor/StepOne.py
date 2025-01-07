from StepBase import StepBase
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook

class StepOne(StepBase):
    def __init__(self):
        super().__init__()
        self.file_path = None
        self.columns = []
        self.selected_columns = []
        self.step = 1

    def setup_widgets(self, parent_frame):
        # Create step label
        self.step_label = ttk.Label(parent_frame, 
                                  text="Step 1: Upload your Hendrix Input Sheet (HIS)", 
                                  font=("Arial", 14))
        self.step_label.pack(pady=10)

        # Create file upload button
        self.upload_btn = ttk.Button(parent_frame, 
                                   text="Upload HIS Excel File",
                                   command=lambda: self.upload_file([("Excel files", "*.xlsx *.xls")]))
        self.upload_btn.pack(pady=10)

        # Create column selection widgets
        self.column_label = ttk.Label(parent_frame, 
                                    text="Select Columns to Keep (including Sequence #)", 
                                    font=("Arial", 14))
        self.column_label.pack(pady=10)

        # Create listbox for column selection
        self.listbox_frame = ttk.Frame(parent_frame)
        self.listbox_frame.pack(pady=10)
        
        self.column_listbox = tk.Listbox(self.listbox_frame, 
                                       selectmode=tk.MULTIPLE, 
                                       font=("Arial", 12), 
                                       width=50, 
                                       height=10)
        self.column_listbox.pack(side=tk.LEFT, fill=tk.BOTH)
        
        self.scrollbar = ttk.Scrollbar(self.listbox_frame, 
                                     orient=tk.VERTICAL, 
                                     command=self.column_listbox.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.column_listbox.config(yscrollcommand=self.scrollbar.set)

        # Add process and skip buttons
        self.process_btn = ttk.Button(parent_frame, 
                                    text="Parse Data and Move to Next Step",
                                    command=self.save_data)
        self.process_btn.pack(pady=10)

        self.skip_btn = ttk.Button(parent_frame, 
                                 text="Skip This Step",
                                 command=self.next_step)
        self.skip_btn.pack(pady=10)

    def process_file(self):
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active
            self.columns = [cell.value for cell in sheet[1]]
            self.column_listbox.delete(0, tk.END)
            for column in self.columns:
                self.column_listbox.insert(tk.END, column)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read file: {e}")

    def save_data(self):
        if not self.file_path:
            messagebox.showerror("Error", "Please upload a file first!")
            return

        selected_indices = self.column_listbox.curselection()
        self.selected_columns = [self.column_listbox.get(i) for i in selected_indices]
        
        if not self.selected_columns:
            messagebox.showerror("Error", "Please select at least one column to keep!")
            return

        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active

            with open("extractHIS_seq_facID_existingTrans_primaryRiser_secondaryRiser.txt", "w") as file:
                file.write("\t".join(self.selected_columns) + "\n")
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:
                        row_data = [str(row[self.columns.index(col)]) for col in self.selected_columns]
                        file.write("\t".join(row_data) + "\n")

            messagebox.showinfo("Success", f"Data from Step {self.step} saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse file: {e}")