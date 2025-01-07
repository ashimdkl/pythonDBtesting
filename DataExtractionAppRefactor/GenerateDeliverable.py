from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re

class GenerateDeliverable:
    def __init__(self):
        self.files_to_combine = [
            "extractHIS_seq_facID_existingTrans_primaryRiser_secondaryRiser.txt",
            "extractFusingCoordination_newOrExistingFusing.txt",
            "extractConstrucStakingReport_framing_type_direction_length.txt",
            "extractPoleType.txt",
            "extractGuyUsage_seq_elementType_usage.txt",
            "extractMAX_sequence_MaxForce.txt"
        ]

    def generate_report(self):
        """Generate the final combined report from all generated txt files"""
        parsed_data = {}
        
        # Parse each input file
        for file_path in self.files_to_combine:
            try:
                with open(file_path, 'r') as file:
                    lines = file.readlines()
                
                if "extractHIS_seq" in file_path:
                    parsed_data['his_seq'] = self._parse_his_seq(lines)
                elif "extractFusingCoordination" in file_path:
                    parsed_data['fusing'] = self._parse_fusing_coordination(lines)
                elif "extractConstrucStakingReport" in file_path:
                    parsed_data['construction'] = self._parse_construction_staking(lines)
                elif "extractPoleType" in file_path:
                    parsed_data['pole_type'] = self._parse_pole_type(lines)
                elif "extractGuyUsage" in file_path:
                    parsed_data['guy_usage'] = self._parse_guy_usage(lines)
                elif "extractMAX_sequence_MaxForce" in file_path:
                    parsed_data['max_force'] = self._parse_max_force(lines)
                    
            except Exception as e:
                messagebox.showerror("Error", f"Failed to read {file_path}: {e}")
                return False

        # Combine all parsed data
        combined_data = self._combine_data(parsed_data)

        # Get save location from user
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if save_path:
            if self._save_to_excel(combined_data, save_path):
                self._add_stringing_data(save_path)
                messagebox.showinfo("Success", "Report generated successfully!")
                return True
        
        return False

    def _parse_his_seq(self, lines):
        """Parse HIS data from text file"""
        data = {}
        pattern = r"(\d{4})\s+([\d\.]+|None)\s+(\d+|None)\s+(Replace|None)?\s+(Replace|None)?"
        
        for line in lines[1:]:  # Skip header line
            match = re.match(pattern, line.strip())
            if match:
                seq, fac_id, existing_trans, primary_riser, secondary_riser = match.groups()
                data[seq] = {
                    'facility_id': fac_id,
                    'existing_transformers': existing_trans,
                    'primary_riser': primary_riser,
                    'secondary_riser': secondary_riser
                }
        
        return data

    def _parse_fusing_coordination(self, lines):
        """Parse fusing coordination data"""
        data = {}
        pattern = r"(\d{4})\s+(.+)"
        
        for line in lines[1:]:  # Skip header line
            match = re.match(pattern, line.strip())
            if match:
                seq, existing = match.groups()
                if seq not in data:
                    data[seq] = []
                data[seq].append(existing)
        
        return data

    def _parse_construction_staking(self, lines):
        """Parse construction staking data"""
        data = {}
        
        for line in lines[2:]:  # Skip header and separator
            parts = [part.strip() for part in line.split('|')]
            if len(parts) == 7:
                seq, type_, lat, lon, framing, anchor_dir, lead_length = parts
                if seq not in data:
                    data[seq] = []
                    
                data[seq].append({
                    'type': type_,
                    'latitude': lat,
                    'longitude': lon,
                    'framing': framing,
                    'anchor_direction': anchor_dir,
                    'lead_length': lead_length
                })
            
        return data

    def _parse_pole_type(self, lines):
        """Parse pole type data"""
        data = {}
        pattern = r"(\d{4})\s+([\w\-\.]+)"
        
        for line in lines[1:]:  # Skip header line
            match = re.match(pattern, line.strip())
            if match:
                seq, pole_type = match.groups()
                data[seq] = pole_type
                
        return data

    def _parse_guy_usage(self, lines):
        """Parse guy usage data"""
        data = {}
        
        for line in lines[2:]:  # Skip header and separator
            parts = [part.strip() for part in line.split('|')]
            if len(parts) == 4:
                seq, element_label, element_type, max_usage = parts
                if seq not in data:
                    data[seq] = []
                    
                data[seq].append({
                    'element_label': element_label,
                    'element_type': element_type,
                    'max_usage': max_usage
                })
            
        return data

    def _parse_max_force(self, lines):
        """Parse maximum force data"""
        data = {}
        
        for line in lines[2:]:  # Skip header and separator
            parts = [part.strip() for part in line.split('|')]
            if len(parts) == 3:
                seq, max_force, soil_class = parts
                data[seq] = {
                    'max_force': max_force,
                    'soil_class': soil_class
                }
            
        return data

    def _combine_data(self, parsed_data):
        """Combine all parsed data into single structure"""
        combined = {}
        
        # Get all unique sequence numbers
        all_sequences = set()
        for key in ['his_seq', 'fusing', 'construction', 'pole_type', 'guy_usage', 'max_force']:
            all_sequences.update(parsed_data[key].keys())

        # Combine data for each sequence
        for seq in all_sequences:
            combined[seq] = {
                'facility_id': parsed_data['his_seq'].get(seq, {}).get('facility_id', ''),
                'existing_transformers': parsed_data['his_seq'].get(seq, {}).get('existing_transformers', ''),
                'primary_riser': parsed_data['his_seq'].get(seq, {}).get('primary_riser', ''),
                'secondary_riser': parsed_data['his_seq'].get(seq, {}).get('secondary_riser', ''),
                'existing_or_new_tap': parsed_data['fusing'].get(seq, []),
                'construction': parsed_data['construction'].get(seq, []),
                'pole_type': parsed_data['pole_type'].get(seq, ''),
                'guy_usage': parsed_data['guy_usage'].get(seq, []),
                'max_force': parsed_data['max_force'].get(seq, {}).get('max_force', ''),
                'soil_class': parsed_data['max_force'].get(seq, {}).get('soil_class', '')
            }

        return combined

    def _save_to_excel(self, data, file_path):
        """Save combined data to Excel with formatting"""
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data Report"

            # Add headers
            headers = [
                'sequence', 'facility_id', 'existing_transformers', 'primary_riser',
                'secondary_riser', 'existing_or_new_tap', 'type', 'latitude', 'longitude',
                'framing', 'anchor_direction', 'lead_length', 'pole_type', 'element_label',
                'element_type', 'max_usage', 'max_force', 'soil_class', 'description'
            ]
            sheet.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)
            for cell in sheet["1:1"]:
                cell.fill = header_fill
                cell.font = header_font

            # Add data with alternating colors
            light_green = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            light_blue = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            current_fill = light_green
            previous_seq = None

            # Add data rows
            for seq, info in sorted(data.items(), key=lambda x: int(re.findall(r'\d+', x[0])[0])):
                max_rows = max(
                    len(info['existing_or_new_tap']),
                    len(info['construction']),
                    len(info['guy_usage']),
                    1
                )

                for i in range(max_rows):
                    row = []
                    if seq != previous_seq:
                        row = [
                            seq,
                            info['facility_id'],
                            info['existing_transformers'],
                            info['primary_riser'],
                            info['secondary_riser']
                        ]
                        current_fill = light_blue if current_fill == light_green else light_green
                        previous_seq = seq
                    else:
                        row = [''] * 5

                    # Add tap info
                    row.append(info['existing_or_new_tap'][i] if i < len(info['existing_or_new_tap']) else '')

                    # Add construction data
                    if i < len(info['construction']):
                        const = info['construction'][i]
                        row.extend([
                            const['type'],
                            const['latitude'],
                            const['longitude'],
                            const['framing'],
                            const['anchor_direction'],
                            const['lead_length']
                        ])
                    else:
                        row.extend([''] * 6)

                    # Add pole type
                    row.append(info['pole_type'] if i == 0 else '')

                    # Add guy usage data
                    if i < len(info['guy_usage']):
                        guy = info['guy_usage'][i]
                        row.extend([
                            guy['element_label'],
                            guy['element_type'],
                            guy['max_usage']
                        ])
                    else:
                        row.extend([''] * 3)

                    # Add max force and soil class
                    row.append(info['max_force'] if i == 0 else '')
                    row.append(info['soil_class'] if i == 0 else '')
                    row.append('')  # Empty description column

                    # Append row and style it
                    sheet.append(row)
                    for cell in sheet[sheet.max_row]:
                        cell.fill = current_fill

            # Adjust column widths
            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            workbook.save(file_path)
            return True

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {e}")
            return False

    def _add_stringing_data(self, file_path):
        """Add stringing chart data to the Excel file"""
        try:
            # Read stringing chart data
            neutral_data = self._parse_stringing_file(
                "extractStringingChartNeutralSpan_section_seq_totalSpanLength_circuitType.txt",
                is_primary=False
            )
            primary_data = self._parse_stringing_file(
                "extractStringingChartPrimary_section_struct_circuitType_spanLength_total.txt",
                is_primary=True
            )

            workbook = load_workbook(file_path)

            # Create and populate Primary Stringing sheet
            primary_sheet = workbook.create_sheet("Primary Stringing Data")
            primary_headers = [
                "Section #", "Structure -> Structure", "Circuit Type",
                "Circuit Value", "Span Length", "Result", "Sequences"
            ]
            primary_sheet.append(primary_headers)
            for row in primary_data:
                primary_sheet.append(row)

            # Create and populate Neutral Span sheet
            neutral_sheet = workbook.create_sheet("Neutral Span Stringing Data")
            neutral_headers = [
                "Section #", "Sequence #s", "Total Span Length", "Circuit Type"
            ]
            neutral_sheet.append(neutral_headers)
            for row in neutral_data:
                neutral_sheet.append(row)

            # Style headers in both sheets
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)
            
            for sheet in [primary_sheet, neutral_sheet]:
                for cell in sheet["1:1"]:
                    cell.fill = header_fill
                    cell.font = header_font

            # Adjust column widths
            for sheet in [primary_sheet, neutral_sheet]:
                for column in sheet.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            workbook.save(file_path)
            return True

        except Exception as e:
            messagebox.showerror("Error", f"Failed to add stringing data: {e}")
            return False

    def _parse_stringing_file(self, file_path, is_primary):
        """Parse stringing chart data from file"""
        data = []
        with open(file_path, 'r') as file:
            lines = file.readlines()[2:]  # Skip header and separator line
            for line in lines:
                parts = [part.strip() for part in line.split('|')]
                if is_primary:
                    data.append(parts)
                else:
                    data.append(parts)
                    
        return data

    def format_cell_value(self, value):
        """Format cell value for Excel"""
        try:
            if isinstance(value, (int, float)):
                return f"{value:.2f}"
            return str(value) if value is not None else ""
        except:
            return ""

    def apply_styles(self, cell, fill=None, font=None, alignment=None):
        """Apply Excel cell styles"""
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment

    def get_formatted_row(self, data, widths):
        """Format row data with consistent widths"""
        return [f"{str(val):<{widths.get(idx, 10)}}" for idx, val in enumerate(data)]

    def set_column_width(self, sheet, column, width):
        """Set Excel column width"""
        sheet.column_dimensions[get_column_letter(column)].width = width