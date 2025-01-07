import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from ttkthemes import ThemedTk
from openpyxl import load_workbook
from StepOne import StepOne
from StepTwo import StepTwo
from StepThree import StepThree
from StepFour import StepFour
from StepFive import StepFive
from StepSix import StepSix
from StepSeven import StepSeven
from GenerateDeliverable import GenerateDeliverable

class DataExtractionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Extraction App")
        self.root.geometry('1920x1080')
        
        self.step_handlers = {
            1: StepOne(),
            2: StepTwo(),
            3: StepThree(),
            4: StepFour(),
            5: StepFive(),
            6: StepSix(),
            7: StepSeven(),
        }
        
        # Inject app reference into each step handler
        for handler in self.step_handlers.values():
            handler.app = self
            
        self.generate_deliverable = GenerateDeliverable()
        self.current_step = 1
        
        # Additional frames
        self.output_frame = None
        
        self.setup_gui()

    def setup_gui(self):
        # Setup intro frame
        self.intro_frame = ttk.Frame(self.root)
        self.intro_frame.pack(expand=True, fill="both")
        self.start_btn = ttk.Button(self.intro_frame, text="Click to Begin Analysis!", 
                                  command=self.start_analysis)
        self.start_btn.pack(pady=20)

        # Setup main and output frames
        self.main_frame = ttk.Frame(self.root)
        self.output_frame = ttk.Frame(self.root)
        self.output_frame.pack(expand=True, fill="both")
        
        # Initialize first step
        self.setup_step_widgets()

    def setup_step_widgets(self):
        # Get current step handler
        step_handler = self.step_handlers[self.current_step]
        
        # Setup widgets for current step
        step_handler.setup_widgets(self.main_frame)
        
        # Setup navigation frame and buttons
        self.setup_navigation_buttons()

    def setup_navigation_buttons(self):
        # Create navigation frame
        nav_frame = ttk.Frame(self.main_frame)
        nav_frame.pack(pady=10)
        
        if self.current_step < 7:
            # Add Next Step button
            next_btn = ttk.Button(nav_frame, text="Next Step", 
                                command=self.next_step)
            next_btn.pack(side=tk.RIGHT, padx=5)
            
            # Add Skip Step button if not last step
            skip_btn = ttk.Button(nav_frame, text="Skip This Step", 
                                command=self.next_step)
            skip_btn.pack(side=tk.RIGHT, padx=5)
            
        else:
            # For step 7, show Generate Report button
            gen_report_btn = ttk.Button(nav_frame, text="Generate Report", 
                                      command=self.generate_report)
            gen_report_btn.pack(side=tk.RIGHT, padx=5)

        # Add Previous Step button if not first step
        if self.current_step > 1:
            prev_btn = ttk.Button(nav_frame, text="Previous Step", 
                                command=self.previous_step)
            prev_btn.pack(side=tk.LEFT, padx=5)

    def start_analysis(self):
        """Start the analysis process"""
        self.intro_frame.pack_forget()
        self.main_frame.pack(expand=True, fill="both")
        self.current_step = 1
        self.setup_step_widgets()

    def next_step(self):
        """Navigate to next step"""
        if self.current_step < 7:
            # Save current step data
            self.step_handlers[self.current_step].save_data()
            
            # Clear current widgets
            for widget in self.main_frame.winfo_children():
                widget.destroy()
                
            # Move to next step
            self.current_step += 1
            self.setup_step_widgets()

    def previous_step(self):
        """Navigate to previous step"""
        if self.current_step > 1:
            # Clear current widgets
            for widget in self.main_frame.winfo_children():
                widget.destroy()
                
            # Move to previous step
            self.current_step -= 1
            self.setup_step_widgets()

    def generate_report(self):
        """Generate final report"""
        # Save current step data first
        self.step_handlers[self.current_step].save_data()
        # Generate report using collected data
        self.generate_deliverable.generate_report()

    def go_to_step8(self):
        """Setup Step 8 interface"""
        self.main_frame.pack_forget()
        self.step8_frame = ttk.Frame(self.root)
        self.step8_frame.pack(expand=True, fill="both")
        self.step8_label = ttk.Label(self.step8_frame, 
                                   text="Step 8: Generate Guy Calc Report Deliverable",
                                   font=("Arial", 14))
        self.step8_label.pack(pady=10)
        
        self.upload_btn_8 = ttk.Button(self.step8_frame,
                                     text="Please Select Your Generated Report",
                                     command=self.upload_generated_report)
        self.upload_btn_8.pack(pady=10)

    def upload_generated_report(self):
        """Handle generated report upload"""
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                workbook = load_workbook(file_path)
                if "Data Report" in workbook.sheetnames:
                    self.generate_deliverable.parse_guy_calc_report(file_path)
                else:
                    messagebox.showerror("Error", "Worksheet 'Data Report' does not exist.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to parse file: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = DataExtractionApp(root)
    root.mainloop()